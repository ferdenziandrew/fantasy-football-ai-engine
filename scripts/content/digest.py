"""
Weekly research digest (Phase 3): pulls recent player news from two RSS feeds, matches
items to players in our own database, and asks Claude to summarize the fantasy-relevant
takeaway and flag a direction (UP/DOWN/NEUTRAL/PENDING) plus a separate news tone for
each match.

Deliberately does NOT write to rankings.blurb automatically -- outputs a plain markdown
report to content/drafts/ for Andrew to review and manually decide what becomes a real
Note update. This matches the design already settled in ROADMAP.md: news feeds into
Notes as a human-approved layer, not a silent auto-override of Andrew-authored content.

Two sources, added together 2026-08-24 to catch up on missed preseason news faster
(both are small-rolling-window feeds, not deep archives -- so more sources = more
coverage per run, not a substitute for running this regularly):

- RotoWire (confirmed reachable/clean): title format is "Player Name: headline", a
  clean, structured split. <ttl>10</ttl> -- built for frequent polling, exposes only a
  handful of the very latest items at any moment.
- RotoBaller (`/player-news/feed`, NOT the general `/feed`, which is long-form articles,
  and NOT `/nfl/feed`, which is a dead WordPress comments feed that returns nothing --
  confirmed 2026-08-24 via direct inspection): cross-sport feed, filtered here to items
  tagged "Fantasy Football & NFL". Titles are natural-sentence headlines with the player
  name woven in ("George Kittle Still on PUP List") -- no clean separator to split on,
  so matching here works by searching for a KNOWN player's normalized full name inside
  the normalized title text instead (see match_rotoballer_items). This is a looser,
  substring-based heuristic (vs. RotoWire's exact structured match) -- full first+last
  name combinations are distinctive enough to make false positives unlikely in
  practice, but it's a real trade-off worth knowing about if an odd match shows up.
  pubDate is unreliable on this feed (empty on every item but the first one observed),
  so it's carried through for display only, not relied on for sorting/filtering.

Matching is deliberately against the FULL players table, not just players already in
our rankings table -- a player excluded from scoring (e.g. by scoring.py's
min_games_to_score filter) can still be exactly the kind of situational riser this
digest exists to catch (see the MarShawn Lloyd case logged in ROADMAP.md's handcuff-
rankings note).

v3 (2026-08-24, Andrew's ask): a persistent per-player history
(data/processed/digest_history.json) is now read before summarizing and written after,
so the digest can tell "is this genuinely new information, or an expected continuation
of something already known" -- without it, a routine practice-absence for an
already-reported injury (Ashton Jeanty) read identically to fresh bad news. Design:
  - Each player's last few entries (date, impact, tone, summary) are fed back into the
    Claude prompt, with explicit instruction to treat a mere continuation as NEUTRAL,
    not re-trigger the same UP/DOWN.
  - A LONG_TERM flag (season-ending injury / IR / 4+ week absence) exempts a player's
    history from the normal 3-week aging-out window -- Andrew's catch: a flat time-based
    expiry would forget an out-for-season player and treat his eventual return as a
    surprise. The flag is re-derived from the LATEST classification each run, so it
    clears itself automatically once news no longer indicates long-term status (e.g. he
    returns to practice) -- no separate reset logic needed.
  - PENDING items are reminded ONCE PER CALENDAR DAY (tracked via
    last_pending_alert_date), not on every run, since Andrew's planned run cadence
    (~3x/day now, possibly ~every 10 min later) would otherwise repeat the same
    reminder many times in one day.
  - The report only gets WRITTEN if something actually changed (a player's impact/tone
    differs from their last known status, or they're new) or there's a same-day-unseen
    pending reminder -- a run that finds nothing new writes nothing, per Andrew's
    "notify me on changes" ask, rather than a repetitive file every run.

v4 (2026-08-25, real-run calibration): two prompt refinements from a second live run --
a veteran starter resting in a PRESEASON GAME (not practice, not injury-related) is
routine and should read NEUTRAL, not DOWN (caught on Josh Allen); an undiagnosed
"being looked at by trainers"-type report should get PENDING, not a premature DOWN,
since the actual severity isn't known yet (caught on Brian Thomas Jr.). Not addressed
here, logged in ROADMAP.md instead as its own feature: cross-player ripple effects
(e.g. Kirk Cousins being the Week 1 starter should inform how to read news about his
own pass-catchers; Ashton Jeanty's injury should flag Mike Washington's snaps as more
relevant) -- this needs real depth-chart/relationship data the digest doesn't have
yet, not a prompt tweak.

v5 (2026-08-25, same-day follow-up): two more refinements from a third live run --
(1) the PRESEASON-GAME guidance from v4 was too blunt (blanket "not a signal"). Andrew's
counter-example: JaDarian Price being deliberately shut down for the last two preseason
games with "we've seen enough" framing IS a real signal (usually UP -- it means the
team has already decided in his favor), unlike a settled veteran starter just resting.
The real test is WHY a player didn't play, not whether it was a preseason game.
(2) `append_validation_log()` writes every reported entry to a persistent CSV
(data/processed/digest_validation_log.csv, one growing file, not one per day) with
blank `correct` and `andrew_notes` columns -- built because Andrew was validating
classifications by narrating them back in chat one at a time, which doesn't scale and
wasn't captured anywhere durable. This becomes a real labeled dataset over time.
Also: the terminal's "Summarizing" line now shows (TEAM-POSITION) per player, since a
bare name wasn't always enough to place who was being discussed.

v6 (2026-08-25, same-day follow-up #2): Andrew's ask -- he wants final say on the
IMPACT/TONE classification, not just a passive log of what the AI decided (right now a
wrong AI call would silently keep feeding "prior status" to future runs uncorrected).
Added two more columns to the validation log -- `override_impact` and `override_tone`
-- and a companion script, apply_validation_overrides.py, that patches
digest_history.json to match Andrew's correction once filled in. The AI still proposes
fast (handles the volume), but Andrew's correction is what actually sticks and feeds
forward, not the AI's original call, once he's reviewed it.

v7 (2026-08-25, first real override review): Andrew's first full validation pass
corrected 7 of 12 overridden entries from the SAME pattern -- a single positive camp
report (Gainwell, Skattebo, Sarratt, Williams, Adams, Hurst, Tuten) had been flagged UP
off one good report, when it should stay NEUTRAL until corroborated. Added explicit
guidance: don't flag UP off a single camp/practice report, wait for a 2nd/3rd
consecutive positive report or something more concrete (depth chart listing, real
usage numbers, a coach naming a starter, a contract/trade).

Run manually for now -- no scheduling yet, per the project's automation-timing rule
(hold off until output quality is validated over a few real runs). Two daily reminders
(11:50am, 5:30pm) are scheduled as of 2026-08-25 to build the habit -- reminders only,
not automated runs, per the same rule.

Usage:
    py scripts/content/digest.py
"""

import json
import re
import sqlite3
import sys
import time
import xml.etree.ElementTree as ET
from datetime import date, datetime, timezone
from pathlib import Path

import requests
from anthropic import Anthropic
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "data" / "db" / "fantasy_football.db"
OUTPUT_DIR = PROJECT_ROOT / "content" / "drafts"
HISTORY_PATH = PROJECT_ROOT / "data" / "processed" / "digest_history.json"
VALIDATION_LOG_PATH = PROJECT_ROOT / "data" / "processed" / "digest_validation_log.xlsx"
# v17 (2026-09-03, Andrew's ask after a real Excel-lock crash lost 5 rows): if
# wb.save() below can't get the file (Andrew has it open), the run's regular rows
# are queued here instead of lost -- see the v17 note in append_validation_log's
# docstring for the full design.
VALIDATION_LOG_PENDING_PATH = PROJECT_ROOT / "data" / "processed" / "digest_validation_log_pending.json"

# Reuse the ingest layer's proven name-matching logic rather than duplicating it.
sys.path.insert(0, str(PROJECT_ROOT / "scripts" / "ingest"))
from ffc import normalize_name  # noqa: E402

ROTOWIRE_FEED_URL = "https://www.rotowire.com/rss/news.php?sport=NFL"
ROTOBALLER_FEED_URL = "https://www.rotoballer.com/player-news/feed"
ROTOBALLER_CATEGORY_MARKER = "fantasy football"  # case-insensitive substring match against item categories
CLAUDE_MODEL = "claude-haiku-4-5-20251001"  # fast/cheap -- lightweight summarization, not deep analysis

# IMPACT (does this change fantasy value) and TONE (how the story itself is framed) are
# two separate flags -- e.g. Justin Jefferson being excited about his new QB is clearly
# POSITIVE tone but NEUTRAL impact (he's elite regardless). PENDING is a real 4th impact
# value, not a NEUTRAL synonym -- for news that's genuinely too early to call.
IMPACT_ORDER = {"UP": 0, "DOWN": 1, "PENDING": 2, "NEUTRAL": 3}
IMPACT_ARROWS = {"UP": "^", "DOWN": "v", "PENDING": "?", "NEUTRAL": "-"}

HISTORY_MAX_AGE_DAYS = 21  # 3 weeks -- Andrew's call, 2026-08-24
HISTORY_ENTRY_CAP = 5
LONG_TERM_ENTRY_CAP = 10  # more room for players we're actively watching through a long absence


def fetch_rotowire_items() -> list[dict]:
    """RotoWire's title format is "Player Name: headline" -- items that don't match
    that shape (rare -- e.g. team-level news with no single player) are skipped rather
    than guessed at."""
    resp = requests.get(ROTOWIRE_FEED_URL, timeout=30)
    resp.raise_for_status()
    root = ET.fromstring(resp.content)

    items = []
    for item in root.findall(".//item"):
        title = (item.findtext("title") or "").strip()
        if ":" not in title:
            continue
        player_name, headline = title.split(":", 1)
        items.append({
            "player_name": player_name.strip(),
            "headline": headline.strip(),
            "description": (item.findtext("description") or "").strip(),
            "link": (item.findtext("link") or "").strip(),
            "pub_date": (item.findtext("pubDate") or "").strip(),
        })
    return items


def fetch_rotoballer_items() -> list[dict]:
    """Cross-sport feed -- filtered here to items with a football-relevant category tag
    (see module docstring for why /player-news/feed specifically, not /feed or
    /nfl/feed). No player-name field or clean title separator, so player_name isn't
    extracted here -- match_rotoballer_items() searches the raw headline text directly
    instead."""
    resp = requests.get(ROTOBALLER_FEED_URL, timeout=30)
    resp.raise_for_status()
    root = ET.fromstring(resp.content)

    items = []
    for item in root.findall(".//item"):
        categories = [c.text or "" for c in item.findall("category")]
        if not any(ROTOBALLER_CATEGORY_MARKER in c.lower() for c in categories):
            continue
        items.append({
            "headline": (item.findtext("title") or "").strip(),
            "description": (item.findtext("description") or "").strip(),
            "link": (item.findtext("link") or "").strip(),
            "pub_date": (item.findtext("pubDate") or "").strip(),
        })
    return items


def build_name_lookup(conn: sqlite3.Connection) -> dict:
    """normalized_name -> [(gsis_id, full_name, position, current_team), ...]. Matches
    against the FULL players table, not just currently-ranked players -- see module
    docstring. current_team included so the report can show it."""
    rows = conn.execute("SELECT gsis_id, full_name, position, current_team FROM players").fetchall()
    lookup: dict = {}
    for gsis_id, full_name, position, current_team in rows:
        key = normalize_name(full_name)
        lookup.setdefault(key, []).append((gsis_id, full_name, position, current_team))
    return lookup


def match_rotowire_items(items: list[dict], name_lookup: dict) -> list[dict]:
    matched = []
    unmatched_count = 0
    for item in items:
        key = normalize_name(item["player_name"])
        candidates = name_lookup.get(key)
        if not candidates:
            unmatched_count += 1
            continue
        if len(candidates) > 1:
            # Ambiguous (rare -- two same-named players). Take the first rather than
            # guess further, same pragmatic call ffc.py makes for genuine ambiguity.
            candidates = candidates[:1]
        gsis_id, full_name, position, current_team = candidates[0]
        matched.append({**item, "gsis_id": gsis_id, "full_name": full_name, "position": position,
                         "team": current_team, "source": "rotowire"})

    print(f"  RotoWire: {len(matched)} matched, {unmatched_count} unmatched "
          f"(coaches, team-level news, non-fantasy-relevant players -- expected, not an error)")
    return matched


def match_rotoballer_items(items: list[dict], name_lookup: dict) -> list[dict]:
    """No clean player-name field on this feed (see module docstring) -- searches for
    each known player's normalized full name as a substring of the normalized headline
    instead. Only checks the headline, not the longer description -- headlines
    consistently lead with the player's name in the examples inspected, and searching
    the full description too would raise the false-positive risk (a name mentioned in
    passing, not the item's actual subject) for little added coverage."""
    matched = []
    unmatched_count = 0
    for item in items:
        normalized_headline = normalize_name(item["headline"])
        found = None
        for normalized_name, candidates in name_lookup.items():
            if normalized_name and normalized_name in normalized_headline:
                found = candidates[0]  # same ambiguity call as RotoWire's matcher
                break
        if not found:
            unmatched_count += 1
            continue
        gsis_id, full_name, position, current_team = found
        matched.append({**item, "gsis_id": gsis_id, "full_name": full_name, "position": position,
                         "team": current_team, "source": "rotoballer"})

    print(f"  RotoBaller: {len(matched)} matched, {unmatched_count} unmatched "
          f"(coaches, team-level news, non-fantasy-relevant players -- expected, not an error)")
    return matched


def group_by_player(matched: list[dict]) -> list[dict]:
    """Merges multiple items about the SAME player (often the same underlying event
    reported by both outlets) into one group, so Claude sees all of it and writes ONE
    synthesized summary instead of near-identical duplicate entries. Order in the input
    list is preserved for the first occurrence of each player."""
    groups: dict = {}
    order: list = []
    for item in matched:
        gsis_id = item["gsis_id"]
        if gsis_id not in groups:
            groups[gsis_id] = {
                "gsis_id": gsis_id, "full_name": item["full_name"], "position": item["position"],
                "team": item["team"], "items": [],
            }
            order.append(gsis_id)
        groups[gsis_id]["items"].append({
            "headline": item["headline"], "description": item["description"],
            "link": item["link"], "source": item["source"],
        })
    return [groups[gsis_id] for gsis_id in order]


def load_history() -> dict:
    if not HISTORY_PATH.exists():
        return {}
    return json.loads(HISTORY_PATH.read_text())


def save_history(history: dict) -> None:
    HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    HISTORY_PATH.write_text(json.dumps(history, indent=2))


def prune_history(history: dict, today: date) -> dict:
    """Drops a player entirely once their most recent entry is older than
    HISTORY_MAX_AGE_DAYS -- UNLESS their latest entry was flagged long_term (season-
    ending injury/IR/4+ week absence), in which case they're kept indefinitely (with a
    larger entry cap) until a future update's classification clears the flag on its
    own. See module docstring for why a flat time-based expiry isn't enough here."""
    pruned = {}
    for gsis_id, record in history.items():
        entries = record.get("entries", [])
        if not entries:
            continue
        is_long_term = entries[-1].get("long_term", False)
        if not is_long_term:
            last_date = date.fromisoformat(entries[-1]["date"])
            if (today - last_date).days > HISTORY_MAX_AGE_DAYS:
                continue
        cap = LONG_TERM_ENTRY_CAP if is_long_term else HISTORY_ENTRY_CAP
        record["entries"] = entries[-cap:]
        pruned[gsis_id] = record
    return pruned


def format_history_for_prompt(prior_entries: list[dict]) -> str:
    if not prior_entries:
        return "No prior history for this player -- this is the first time we've seen news about them."
    lines = ["Prior known status (most recent last):"]
    for e in prior_entries:
        lines.append(f"  {e['date']}: impact={e['impact']}, tone={e['tone']} -- {e['summary']}")
    return "\n".join(lines)


def summarize_with_claude(client: Anthropic, group: dict, prior_entries: list[dict]) -> dict:
    """One call per PLAYER (a group may bundle multiple items from different sources
    reporting the same story), not one call per raw item. Now also passes prior history
    so Claude can judge whether this is genuinely new or an expected continuation."""
    items_text = "\n\n".join(
        f"[{i}] ({it['source']}) {it['headline']}\n{it['description']}"
        for i, it in enumerate(group["items"], start=1)
    )
    multiple_note = (
        " These are multiple reports of the same underlying story -- synthesize them "
        "into ONE summary, don't just restate one and ignore the rest."
        if len(group["items"]) > 1 else ""
    )
    history_text = format_history_for_prompt(prior_entries)

    prompt = (
        f"You're a fantasy football analyst. Given the following news about {group['full_name']} "
        f"({group['position']}, {group['team'] or 'no current team'}), write:\n"
        f"1. A one-sentence fantasy-relevant summary (what a fantasy manager needs to know).{multiple_note}\n"
        f"2. A news tone flag: POSITIVE, NEGATIVE, or NEUTRAL -- how the story itself is framed, "
        f"independent of whether it actually changes his fantasy value.\n"
        f"3. A ranking impact flag: UP, DOWN, NEUTRAL, or PENDING -- does this actually change his "
        f"fantasy value. Positive news doesn't always mean UP (e.g. a good quote about a player who's "
        f"already a clear starter is POSITIVE tone but NEUTRAL impact). A SINGLE positive camp/practice "
        f"report (flashing potential, impressing in a joint practice, one good outing) is NOT enough on its "
        f"own to flag UP -- that's noise until it's corroborated. Check the prior status below: if this is "
        f"the first report of this kind for this player, impact should be NEUTRAL even with POSITIVE tone; "
        f"only flag UP if EITHER of these is true, independently -- neither needs the other: (1) this is "
        f"the 3rd consecutive positive report for this situation with no negative reports in between, or "
        f"(2) something concrete enough that no repetition is needed at all -- a depth chart listing, a "
        f"snap-count/target-share number, a coach or team naming him the new starter, a role opening up to "
        f"him because of a teammate's injury or departure, or a contract/trade. Condition (2) qualifies for "
        f"UP on the very first report, regardless of any prior history for this player. When UP fires "
        f"because of condition (1) rather than (2), say so directly in the summary -- e.g. \"third straight "
        f"positive camp report\" -- so it's clear from the row alone why this is being treated as real "
        f"signal instead of noise. PRESEASON GAME participation needs "
        f"judgment, not a blanket rule -- a clear veteran starter resting in a preseason game is routine, "
        f"NEUTRAL, not a signal either way. But a roster-bubble/competition player being deliberately held "
        f"out with language like 'they've seen enough' or 'earned a roster spot' IS a real signal (usually "
        f"UP -- it implies the team has already decided in his favor) -- the distinguishing question is "
        f"WHY he didn't play: standard rest for someone whose role is already settled = no signal; a "
        f"decision reflecting the team's actual evaluation of a player whose role is still being decided = "
        f"real signal. Use PENDING, not a premature DOWN, when something concerning is reported but "
        f"genuinely undiagnosed yet -- e.g. 'being looked at by trainers' or 'left with an issue' with no "
        f"diagnosis or missed-time estimate given. Don't assume worst case; wait for the actual diagnosis "
        f"before committing to DOWN. IMPORTANT: check the prior status below first -- if this news is just "
        f"an expected continuation of something already known (e.g. a routine practice absence for an "
        f"injury already reported as DOWN), the impact should usually be NEUTRAL, not a repeat of the same "
        f"UP/DOWN. Only flag a new UP/DOWN if something has genuinely changed or escalated since the prior "
        f"status.\n"
        f"4. A long-term flag: YES if this news indicates a season-ending injury, IR placement, or an "
        f"absence of 4+ weeks; NO otherwise.\n"
        f"5. An event-driven flag: YES if this is grounded in something that actually happened or was "
        f"said by someone inside the team -- an injury, a practice/game status change, a roster or "
        f"depth-chart move, a transaction, or a quote from a coach, coordinator, GM, or the player "
        f"himself about his role, health, or outlook (coach/insider commentary counts as actionable even "
        f"without a hard event behind it -- e.g. a coach saying a player looks like his best self this "
        f"time of year is real signal). NO if this is outside media/analyst commentary with no "
        f"organizational quote behind it -- a fantasy analyst's season-outlook or redraft-value opinion "
        f"piece. Default to YES when genuinely unsure -- missing real news is worse than including a "
        f"borderline item.\n\n"
        f"6. A new-information flag: NEW_INFORMATION: YES if this report adds anything materially "
        f"new since the prior status below -- an escalation, a resolution, a genuinely different "
        f"detail, or the first time this kind of news has come up. NO if this is substantially the "
        f"same underlying event as the prior status, even if it's worded very differently, comes "
        f"from a different source, or has more/less detail -- judge by substance, not exact "
        f"wording. A report describing the same practice session, workout, game, or status update "
        f"the prior entry already covers is NO even if the wording is only ~90% different or none "
        f"of it overlaps at all. Default to YES when genuinely unsure, same principle as the "
        f"event-driven flag above -- missing a real update is worse than one redundant entry. This "
        f"is a TRIAL field (added 2026-09-02) -- logged and shown on the sheet but not yet used to "
        f"suppress anything.\n\n"
        f"{history_text}\n\n"
        f"News:\n{items_text}\n\n"
        f"Respond in exactly this format:\n"
        f"SUMMARY: <one sentence>\n"
        f"TONE: <POSITIVE, NEGATIVE, or NEUTRAL>\n"
        f"IMPACT: <UP, DOWN, NEUTRAL, or PENDING>\n"
        f"LONG_TERM: <YES or NO>\n"
        f"EVENT_DRIVEN: <YES or NO>\n"
        f"NEW_INFORMATION: <YES or NO>"
    )

    response = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=250,
        messages=[{"role": "user", "content": prompt}],
    )
    return parse_claude_response(group, response.content[0].text)


def parse_claude_response(group: dict, text: str) -> dict:
    """Split out so the parsing logic can be tested without a real API call."""
    summary, tone, impact, long_term = "", "NEUTRAL", "NEUTRAL", False
    # Lenient default (2026-08-27, Andrew's ask): an unparseable/missing EVENT_DRIVEN
    # line should never silently demote a real item into the opinion section -- only an
    # explicit "NO" does that.
    event_driven = True
    # NEW_INFORMATION (2026-09-02, trial field): same lenient-default philosophy --
    # missing/malformed defaults to YES (treat as new) so nothing is ever silently
    # hidden by a parsing gap. Not yet used to suppress rows (see append_validation_log
    # docstring) -- logged on the sheet only, pending Andrew's ~1-week trust review.
    new_information = True
    for line in text.splitlines():
        if line.startswith("SUMMARY:"):
            summary = line[len("SUMMARY:"):].strip()
        elif line.startswith("TONE:"):
            tone = line[len("TONE:"):].strip().upper()
        elif line.startswith("IMPACT:"):
            impact = line[len("IMPACT:"):].strip().upper()
        elif line.startswith("LONG_TERM:"):
            long_term = line[len("LONG_TERM:"):].strip().upper() == "YES"
        elif line.startswith("EVENT_DRIVEN:"):
            event_driven = line[len("EVENT_DRIVEN:"):].strip().upper() != "NO"
        elif line.startswith("NEW_INFORMATION:"):
            new_information = line[len("NEW_INFORMATION:"):].strip().upper() != "NO"
    return {**group, "summary": summary or text.strip(), "tone": tone, "impact": impact,
            "long_term": long_term, "event_driven": event_driven, "new_information": new_information}


BATCH_STATIC_INSTRUCTIONS = (
    "You're a fantasy football analyst. Below are several players, each under its own "
    "'### PLAYER <N>' marker with their name/position/team, prior known status, and today's "
    "news. For EACH player, independently write:\n"
    "1. A one-sentence fantasy-relevant summary (what a fantasy manager needs to know).\n"
    "2. A news tone flag: POSITIVE, NEGATIVE, or NEUTRAL -- how the story itself is framed, "
    "independent of whether it actually changes his fantasy value.\n"
    "3. A ranking impact flag: UP, DOWN, NEUTRAL, or PENDING -- does this actually change his "
    "fantasy value. Positive news doesn't always mean UP (e.g. a good quote about a player who's "
    "already a clear starter is POSITIVE tone but NEUTRAL impact). A SINGLE positive camp/practice "
    "report (flashing potential, impressing in a joint practice, one good outing) is NOT enough on its "
    "own to flag UP -- that's noise until it's corroborated. Check the prior status given: if this is "
    "the first report of this kind for this player, impact should be NEUTRAL even with POSITIVE tone; "
    "only flag UP if EITHER of these is true, independently -- neither needs the other: (1) this is "
    "the 3rd consecutive positive report for this situation with no negative reports in between, or "
    "(2) something concrete enough that no repetition is needed at all -- a depth chart listing, a "
    "snap-count/target-share number, a coach or team naming him the new starter, a role opening up to "
    "him because of a teammate's injury or departure, or a contract/trade. Condition (2) qualifies for "
    "UP on the very first report, regardless of any prior history for this player. When UP fires "
    "because of condition (1) rather than (2), say so directly in the summary -- e.g. \"third straight "
    "positive camp report\" -- so it's clear from the row alone why this is being treated as real "
    "signal instead of noise. PRESEASON GAME participation needs "
    "judgment, not a blanket rule -- a clear veteran starter resting in a preseason game is routine, "
    "NEUTRAL, not a signal either way. But a roster-bubble/competition player being deliberately held "
    "out with language like 'they've seen enough' or 'earned a roster spot' IS a real signal (usually "
    "UP -- it implies the team has already decided in his favor) -- the distinguishing question is "
    "WHY he didn't play: standard rest for someone whose role is already settled = no signal; a "
    "decision reflecting the team's actual evaluation of a player whose role is still being decided = "
    "real signal. Use PENDING, not a premature DOWN, when something concerning is reported but "
    "genuinely undiagnosed yet -- e.g. 'being looked at by trainers' or 'left with an issue' with no "
    "diagnosis or missed-time estimate given. Don't assume worst case; wait for the actual diagnosis "
    "before committing to DOWN. IMPORTANT: check the prior status given first -- if this news is just "
    "an expected continuation of something already known (e.g. a routine practice absence for an "
    "injury already reported as DOWN), the impact should usually be NEUTRAL, not a repeat of the same "
    "UP/DOWN. Only flag a new UP/DOWN if something has genuinely changed or escalated since the prior "
    "status.\n"
    "4. A long-term flag: YES if this news indicates a season-ending injury, IR placement, or an "
    "absence of 4+ weeks; NO otherwise.\n"
    "5. An event-driven flag: YES if this is grounded in something that actually happened or was "
    "said by someone inside the team -- an injury, a practice/game status change, a roster or "
    "depth-chart move, a transaction, or a quote from a coach, coordinator, GM, or the player "
    "himself about his role, health, or outlook (coach/insider commentary counts as actionable even "
    "without a hard event behind it -- e.g. a coach saying a player looks like his best self this "
    "time of year is real signal). NO if this is outside media/analyst commentary with no "
    "organizational quote behind it -- a fantasy analyst's season-outlook or redraft-value opinion "
    "piece. Default to YES when genuinely unsure -- missing real news is worse than including a "
    "borderline item.\n\n"
    "6. A new-information flag: NEW_INFORMATION: YES if this report adds anything materially "
    "new since the prior status given -- an escalation, a resolution, a genuinely different "
    "detail, or the first time this kind of news has come up. NO if this is substantially the "
    "same underlying event as the prior status, even if it's worded very differently, comes "
    "from a different source, or has more/less detail -- judge by substance, not exact "
    "wording. A report describing the same practice session, workout, game, or status update "
    "the prior entry already covers is NO even if the wording is only ~90% different or none "
    "of it overlaps at all. Default to YES when genuinely unsure, same principle as the "
    "event-driven flag above -- missing a real update is worse than one redundant entry. This "
    "is a TRIAL field (added 2026-09-02) -- logged and shown on the sheet but not yet used to "
    "suppress anything.\n\n"
    "Treat each player completely independently -- do not let one player's news influence another "
    "player's classification, even if they're on the same team.\n\n"
    "Respond with exactly one block per player, in the SAME ORDER given below, each starting with "
    "that player's own marker line exactly as given, in this exact format:\n"
    "### PLAYER <N>\n"
    "SUMMARY: <one sentence>\n"
    "TONE: <POSITIVE, NEGATIVE, or NEUTRAL>\n"
    "IMPACT: <UP, DOWN, NEUTRAL, or PENDING>\n"
    "LONG_TERM: <YES or NO>\n"
    "EVENT_DRIVEN: <YES or NO>\n"
    "NEW_INFORMATION: <YES or NO>\n"
)

PLAYER_MARKER_RE = re.compile(r"^###\s*PLAYER\s+(\d+)\s*$", re.MULTILINE)


def _build_player_block(index: int, group: dict, prior_entries: list[dict]) -> str:
    """The per-player portion of a batch prompt -- same content summarize_with_claude
    sends for one player (name/position/team, prior status, news items), labeled with
    an index so the batched response can be split back apart per player."""
    items_text = "\n\n".join(
        f"[{i}] ({it['source']}) {it['headline']}\n{it['description']}"
        for i, it in enumerate(group["items"], start=1)
    )
    multiple_note = (
        "\n(Multiple reports of the same underlying story above -- synthesize them into ONE "
        "summary, don't just restate one and ignore the rest.)"
        if len(group["items"]) > 1 else ""
    )
    history_text = format_history_for_prompt(prior_entries)
    return (
        f"### PLAYER {index}\n"
        f"Player: {group['full_name']} ({group['position']}, {group['team'] or 'no current team'})\n"
        f"{history_text}\n\n"
        f"News:\n{items_text}{multiple_note}"
    )


def parse_batch_response(groups: list[dict], text: str) -> dict[int, dict]:
    """Splits a batched response back into one parsed result per player, keyed by the
    1-based index used in the prompt. Reuses parse_claude_response per block so the same
    lenient field-level defaults apply. Indices missing from the response (bad split, or
    Claude skipped one) are simply absent from the returned dict -- the caller is
    responsible for falling back to an individual call for any that are missing."""
    results: dict[int, dict] = {}
    parts = PLAYER_MARKER_RE.split(text)
    # re.split with a capturing group yields [pre-text, idx1, block1, idx2, block2, ...]
    for i in range(1, len(parts), 2):
        try:
            idx = int(parts[i])
        except ValueError:
            continue
        if 1 <= idx <= len(groups):
            results[idx] = parse_claude_response(groups[idx - 1], parts[i + 1])
    return results


def summarize_batch_with_claude(client: Anthropic, groups: list[dict], history: dict) -> list[dict]:
    """v14 (2026-09-02, Andrew's ask, cost reduction): one Claude call for every matched
    player in the run instead of one call per player -- the instruction block above only
    gets sent once instead of N times, roughly halving cost. (Prompt caching was the
    first idea, but Haiku 4.5 requires a 4,096-token minimum cacheable prefix and this
    instruction block is well under that -- caching would silently do nothing at this
    prompt's size, confirmed against Anthropic's docs before building this instead.)
    Never silently drops a player: falls back to the original one-call-per-player
    summarize_with_claude for any player whose block is missing from the batch response,
    or for everyone in the batch if the response was truncated (max_tokens hit, so we
    can't trust where any block actually ended)."""
    if not groups:
        return []
    prior_entries_list = [history.get(g["gsis_id"], {}).get("entries", []) for g in groups]
    prompt = BATCH_STATIC_INSTRUCTIONS + "\n\n" + "\n\n".join(
        _build_player_block(i, g, pe)
        for i, (g, pe) in enumerate(zip(groups, prior_entries_list), start=1)
    )
    max_tok = min(4096, max(500, 120 * len(groups) + 200))
    response = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=max_tok,
        messages=[{"role": "user", "content": prompt}],
    )
    text = response.content[0].text
    results_by_index: dict[int, dict] = {}
    if response.stop_reason == "max_tokens":
        print(f"  Batch response hit max_tokens ({max_tok}) -- discarding, falling back to "
              f"individual calls for all {len(groups)} player(s)")
    else:
        results_by_index = parse_batch_response(groups, text)

    missing = [i for i in range(1, len(groups) + 1) if i not in results_by_index]
    if missing and results_by_index:
        # Partial miss (not a full-batch discard above) -- name the players so a gap is
        # visible in the run log, not just a silent count.
        missing_names = ", ".join(groups[i - 1]["full_name"] for i in missing)
        print(f"  Batch response missing {len(missing)} player(s) ({missing_names}) -- "
              f"falling back to individual calls for just those")
    for i in missing:
        results_by_index[i] = summarize_with_claude(client, groups[i - 1], prior_entries_list[i - 1])

    return [results_by_index[i] for i in range(1, len(groups) + 1)]


def _headline_signature(items: list[dict]) -> str:
    """Stable fingerprint of the raw source headlines behind a classification -- sorted
    so item order (which can vary run to run even for the identical set) doesn't matter.
    v13 (2026-09-01, Andrew's ask): the same top RSS item can still be the newest thing in
    the feed across two runs close together (e.g. a 7 AM and 1 PM pull with nothing new in
    between). Re-classifying identical text can come back with a slightly different
    impact/tone purely from LLM non-determinism -- caught this directly when two manual
    test runs 44 seconds apart on IDENTICAL headlines produced Malik Nabers as PENDING once
    and NEUTRAL the next, writing a spurious duplicate row for nothing."""
    return "|".join(sorted(it["headline"] for it in items))


def is_change(result: dict, history: dict) -> bool:
    """A player counts as "changed" if this is the first time we've seen them, or their
    impact/tone differs from their most recently recorded status. Used to decide what
    goes in the report -- a run where nothing changed writes nothing (see module
    docstring, Andrew's "notify me on changes" ask)."""
    prior_entries = history.get(result["gsis_id"], {}).get("entries", [])
    if not prior_entries:
        return True
    last = prior_entries[-1]
    # v13: same source headline(s) as last time -- not a real change no matter what impact/
    # tone this pass came back with (see _headline_signature docstring). Old history entries
    # predate this field and have no "headline_signature" key, so this simply never matches
    # for them and falls through to the normal comparison below -- no migration needed.
    if last.get("headline_signature") == _headline_signature(result["items"]):
        return False
    return last["impact"] != result["impact"] or last["tone"] != result["tone"]


def update_history(history: dict, results: list[dict], today_str: str) -> dict:
    """Records every result (not just changes) so last-seen dates stay accurate even
    for unchanged/confirmed statuses."""
    for r in results:
        gsis_id = r["gsis_id"]
        record = history.setdefault(gsis_id, {
            "full_name": r["full_name"], "position": r["position"], "team": r["team"],
            "last_pending_alert_date": None, "entries": [],
        })
        record["full_name"] = r["full_name"]
        record["position"] = r["position"]
        record["team"] = r["team"]
        record["entries"].append({
            "date": today_str, "impact": r["impact"], "tone": r["tone"],
            "summary": r["summary"], "long_term": r.get("long_term", False),
            "event_driven": r.get("event_driven", True),
            "headline_signature": _headline_signature(r["items"]),
        })
        if r["impact"] == "PENDING":
            # v10 (2026-08-28, Andrew's ask -- fewer/less noisy reminders): a player
            # who's freshly PENDING today (whether this is a brand new entry or an
            # updated one) already got surfaced today via the regular "What's new"
            # section -- an immediate reminder the SAME day is a redundant duplicate of
            # information Andrew just saw. Marking today as already-alerted here means
            # get_pending_reminders() won't also fire for them today; reminders now
            # start the day AFTER a player goes (or stays) PENDING with no further news.
            record["last_pending_alert_date"] = today_str
    return history


def get_pending_reminders(history: dict, today_str: str) -> list[dict]:
    """Players whose latest known status is PENDING and haven't already been reminded
    today -- mutates last_pending_alert_date in history as a side effect (caller must
    save_history() afterward). Once per calendar day, not once per run -- see module
    docstring on why (Andrew's planned run cadence would otherwise repeat this many
    times in one day)."""
    reminders = []
    for gsis_id, record in history.items():
        entries = record.get("entries", [])
        if not entries or entries[-1]["impact"] != "PENDING":
            continue
        if record.get("last_pending_alert_date") == today_str:
            continue
        reminders.append({
            "gsis_id": gsis_id, "full_name": record["full_name"], "position": record["position"],
            "team": record["team"], "summary": entries[-1]["summary"], "since_date": entries[-1]["date"],
        })
        record["last_pending_alert_date"] = today_str
    return reminders


def _format_player_blocks(entries: list[dict]) -> list[str]:
    """Shared block formatter (arrow, impact/tone header, summary, source links) for one
    player -- used by both the 'What's new' and 'Analyst takes' sections in write_report
    (added 2026-08-27 alongside EVENT_DRIVEN) so the two sections share formatting
    instead of duplicating it."""
    lines = []
    for r in entries:
        arrow = IMPACT_ARROWS.get(r["impact"], "?")
        team = r["team"] or "no current team"
        long_term_tag = " [long-term]" if r.get("long_term") else ""
        lines.append(f"### [{arrow}] {r['full_name']} ({r['position']}, {team}) -- "
                     f"Impact: {r['impact']} | Tone: {r['tone']}{long_term_tag}")
        lines.append(r["summary"])
        source_links = ", ".join(f"[*{it['headline']}*]({it['link']}) ({it['source']})" for it in r["items"])
        lines.append(f"Sources: {source_links}")
        lines.append("")
    return lines


def write_report(changes: list[dict], pending_reminders: list[dict], date_str: str) -> Path | None:
    if not changes and not pending_reminders:
        return None

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"digest_{date_str}.md"
    if out_path.exists():
        # v12 (2026-08-31, Andrew's ask -- running more than once a day): the first run of
        # a day keeps the plain digest_YYYY-MM-DD.md name (unchanged from before, so a normal
        # single-run day looks exactly like it always has). A second-or-later run the SAME day
        # previously overwrote that file outright, silently losing an earlier run's report --
        # now it gets its own HHMM-suffixed file instead, so nothing from an earlier run today
        # disappears.
        out_path = OUTPUT_DIR / f"digest_{date_str}_{datetime.now().strftime('%H%M')}.md"

    lines = [f"# Research Digest -- {date_str}", ""]

    if changes:
        changes = sorted(changes, key=lambda r: IMPACT_ORDER.get(r["impact"], 4))
        # v9 (2026-08-27, Andrew's ask): split items with no real event or team-insider
        # quote behind them (a media outlet's season-outlook/redraft-value opinion
        # piece -- e.g. rotoballer's "could follow his career year" pieces) out of
        # "What's new" into their own section below. Demoted, not dropped -- a wrong
        # EVENT_DRIVEN call should never hide real news. See the EVENT_DRIVEN criteria
        # in summarize_with_claude's prompt (a coach/insider quote counts as YES even
        # with no hard event behind it -- Andrew's Ben Johnson example).
        event_changes = [r for r in changes if r.get("event_driven", True)]
        opinion_changes = [r for r in changes if not r.get("event_driven", True)]

        if event_changes:
            source_counts: dict = {}
            for r in event_changes:
                for it in r["items"]:
                    source_counts[it["source"]] = source_counts.get(it["source"], 0) + 1
            source_summary = ", ".join(f"{count} from {source}" for source, count in source_counts.items())
            lines.append(f"## What's new ({len(event_changes)} players, {sum(source_counts.values())} items: {source_summary})")
            lines.append("")
            lines.extend(_format_player_blocks(event_changes))

        if opinion_changes:
            lines.append(f"## Analyst takes ({len(opinion_changes)} players -- not event-driven, FYI only)")
            lines.append("")
            lines.extend(_format_player_blocks(opinion_changes))

    if pending_reminders:
        lines.append("## Still pending from earlier")
        lines.append("")
        for p in pending_reminders:
            team = p["team"] or "no current team"
            lines.append(f"- **{p['full_name']}** ({p['position']}, {team}) -- "
                         f"pending since {p['since_date']}: {p['summary']}")
        lines.append("")

    out_path.write_text("\n".join(lines))
    return out_path


VALIDATION_LOG_FIELDNAMES = [
    "date", "player", "position", "team", "impact", "tone", "long_term",
    "summary", "headlines", "correct", "andrew_notes", "override_impact", "override_tone",
    "event_driven", "new_information",
]

# Andrew's preferred column widths (2026-08-25 ask, tightened further 2026-08-26 so
# override_impact/override_tone are visible without scrolling) -- re-applied on every
# write so they're always right when he opens the file, not something to redo by hand
# each time. Units are openpyxl's "characters" width, same scale Excel itself uses.
VALIDATION_LOG_COLUMN_WIDTHS = {
    "date": 12, "player": 12, "position": 5, "team": 5, "impact": 9, "tone": 8,
    "long_term": 7, "summary": 20, "headlines": 55, "correct": 6,
    "andrew_notes": 45, "override_impact": 14, "override_tone": 12,
    "event_driven": 8,  # rightmost column (Andrew's ask, 2026-08-27) -- keeps the
    # existing screen-fit layout intact rather than inserting it mid-sheet
    "new_information": 8,  # TRIAL column, 2026-09-02 -- Andrew's ask is to revisit in
    # ~1 week whether this (and event_driven/long_term) are trusted enough to stop
    # needing a visible column at all; see append_validation_log's NEW_INFORMATION note
}


def append_validation_log(changes: list[dict], pending_reminders: list[dict], date_str: str, history: dict) -> int:
    """Appends every reported entry to a persistent Excel workbook
    (data/processed/digest_validation_log.xlsx) for Andrew to review after reading the
    markdown report. Columns beyond the core data are all his to fill in: `correct`
    (Y/N), `andrew_notes` (free text), and `override_impact` / `override_tone` -- if the
    AI's call was wrong, filling in an override here and running
    apply_validation_overrides.py patches that player's actual recorded status in
    digest_history.json, so the CORRECTED value (not the AI's original wrong one) is
    what feeds into future runs' "prior status" context. Andrew's ask 2026-08-25: he
    wants final say on IMPACT/TONE, not just a passive log of what the AI decided.

    v8 (2026-08-25, same-day format switch): originally a plain CSV, but CSV has no way
    to store column widths -- Andrew was manually re-widening the same 4 columns every
    time he opened it, and separately Excel silently reformatted the date column
    (2026-08-25 -> 8/25/2026) on save, which broke apply_validation_overrides.py's date
    matching entirely. Switched to .xlsx (openpyxl, already a project dependency) which
    solves both: column widths persist in the file itself and get re-applied on every
    write, and the date column is written as a real Excel date type + explicit
    YYYY-MM-DD number format instead of a string Excel is free to reinterpret.

    Built 2026-08-25: Andrew was validating classifications by narrating them back in
    chat one at a time, which doesn't scale and isn't captured anywhere durable (the
    history file records what the digest DECIDED, not whether Andrew thought that
    decision was right). This grows into a real labeled dataset over time -- e.g.
    "what fraction of PENDING calls were actually correct" becomes an answerable
    question once there's enough of this, not just a vibe. Appends across runs into
    ONE file (not one file per day) specifically so it accumulates into something
    analyzable, rather than being scattered across many small files.

    v9 (2026-08-27, Andrew's ask -- try it, revert if it doesn't feel right):
    - EVERY write now re-sorts the whole sheet so all "(reminder)" rows sit at the true
      bottom, not just the bottom of that run's own batch (running the digest more than
      once a day was leaving an earlier run's reminders sandwiched between run
      batches). This is a full read-everything/rebuild pass -- existing rows are read
      back by VALUE and rewritten -- so any formatting Andrew has added BEYOND what
      this function manages (fills, date format, column widths are all reapplied
      automatically) will NOT survive, e.g. manually bolding a cell or adding a custom
      color. Worth knowing before deciding whether to keep this.
    - A live "needs review" conditional-formatting highlight on non-reminder rows where
      `correct` is still blank -- tracks Andrew's edits automatically in Excel itself,
      no extra script run needed to clear the highlight once he fills a row in.
    - A hard-stop data validation on override_impact/override_tone that rejects typing
      a value into those cells on a reminder row, since the grey fill alone hasn't
      stopped that mistake from recurring.

    v10 (2026-08-28, Andrew's ask -- pending reminders were "confusing and noisy"):
    get_pending_reminders() still fires once per calendar day per still-PENDING player
    (unchanged -- the daily markdown "Still pending from earlier" section still wants
    that), but this function no longer turns every one of those into a NEW row. A
    reminder row is now upserted per player instead of appended: if that player already
    has an open "(reminder)" row, its summary/date are refreshed in place (Andrew's
    correct/andrew_notes/override_* on that row are left untouched -- his prior
    annotation stays attached to the current row rather than getting orphaned under a
    fresh blank duplicate); only a player with no existing open reminder row gets a
    brand new one. A reminder row's `date` is now the story's since_date (when this
    status was last substantively true) rather than the day the row happened to be
    (re)written, and the redundant "(pending since <date>)" text is dropped from the
    summary since the date column now carries that -- also makes the date column
    filterable/sortable by staleness, which it wasn't before. One-time migration: any
    player with MULTIPLE existing "(reminder)" rows (an artifact of the old
    once-a-day-forever behavior) gets collapsed to one on the first v10 run, keeping
    whichever duplicate has the most of Andrew's own content (correct/andrew_notes/
    override_*) and refreshing its summary/date to the most recent of the group.
    Known limitation, accepted rather than engineered around: matching is by player
    name only, not by story identity -- if a player has a fully-resolved PENDING
    episode and, much later, a brand new unrelated one, the new one would upsert onto
    the old (dormant) reminder row instead of getting a fresh one. Not worth the added
    complexity (a gsis_id + history lookup) for how unlikely that is to happen within a
    single preseason.

    v11 (2026-08-31, Andrew's ask after seeing v10 run for real): two follow-ups once he
    saw actual orphaned rows and a real name collision.
    (1) Auto-clear: a reminder row now requires the player's CURRENT history status to
    still be PENDING or it is dropped entirely on the next write (this needs `history`,
    hence the new parameter -- `pending_reminders` alone can't tell "still pending, not
    re-alerted today" apart from "resolved, no longer pending at all"). Andrew's explicit
    choice over archiving to a separate section or leaving it manual -- he accepted that a
    note/override on a row that resolves is gone with it, not preserved anywhere.
    (2) Collision flag: a fresh regular row (in `changes`) for a player who ALSO still has
    an open reminder row gets a bolded "[Already has an open reminder]" prefix on its
    summary (rich text, via openpyxl's CellRichText/TextBlock) -- Andrew's own wording, kept
    short on purpose since he already knows to use the player-column filter to go find the
    matching grey row, so the prefix doesn't repeat instructions he's already been told.

    v16 (2026-09-02, TRIAL, Andrew's ask -- found via a real Malik Nabers row that got
    re-logged a second time with softened impact instead of being recognized as the same
    underlying practice session already covered): NEW_INFORMATION is a new Claude-judged
    field (see the prompt's rule 6) -- YES if a report adds anything materially new since
    the player's prior status, NO if it's substantially the same event just worded
    differently. It is ONLY logged as a column right now (`new_information`, rightmost) --
    NOT used to suppress a row. Andrew's explicit call: skipping straight to silent
    suppression removes his ability to catch a bad call by looking at the sheet, the same
    way the collision-flag self-reference bug (above) and the v13 duplicate-row bug were
    both caught by him noticing something off in a visible row. Plan: after roughly a
    week of real runs, have a short discussion on whether NEW_INFORMATION's judgment has
    proven reliable enough to actually start suppressing NO rows -- and, at the same time,
    whether `long_term` and `event_driven` (both older, already-trusted columns) have
    proven reliable enough that they no longer need to be visible columns Andrew reviews
    row by row either. Nothing about long_term/event_driven's behavior changes today --
    this is purely a flagged future discussion, not a change.

    v17 (2026-09-03, Andrew's ask -- a real run crashed here because he had the xlsx
    open in Excel at 10am, silently losing that run's 5 rows from the sheet even though
    digest_history.json and the markdown report had already been written): wb.save() is
    now retried a couple of times a few seconds apart (covers a brief transient lock --
    antivirus scan, OneDrive sync -- not the common case of Andrew actually having the
    file open). If it still can't save, this run's `changes` are written to
    VALIDATION_LOG_PENDING_PATH (plain JSON, paired with their own date_str) instead of
    being lost, and the function returns without raising -- main() finishes normally,
    it just logs 0 new rows this run. Every future call to this function (i.e. the next
    scheduled run) checks that file FIRST and folds any queued rows in ahead of its own
    `changes`, using each row's own original date, then clears the queue once a save
    actually succeeds -- so the backlog empties itself out automatically the next time
    the file is closed, with no manual backfill needed. Costs nothing extra (no API
    calls, just local disk I/O) and never blocks -- worst case is rows sit queued for
    however many runs the file stays open, same as this real 5-row backlog did before
    being backfilled by hand.
    """
    import re
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Border, PatternFill, Side
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.formatting.formatting import ConditionalFormattingList
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
    from openpyxl.utils import get_column_letter

    # Light grey fill for "(reminder)" rows -- visually distinct from a real, actionable
    # entry (see the override guardrail below for the stronger fix). A thin border was
    # added 2026-08-28 (Andrew's ask) -- a solid fill on its own hides Excel's default
    # gridlines, making it hard to tell where one column ends and the next begins when
    # scanning across a grey row.
    REMINDER_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
    REMINDER_BORDER_SIDE = Side(style="thin", color="BFBFBF")
    REMINDER_BORDER = Border(left=REMINDER_BORDER_SIDE, right=REMINDER_BORDER_SIDE,
                              top=REMINDER_BORDER_SIDE, bottom=REMINDER_BORDER_SIDE)
    NEEDS_REVIEW_FILL = PatternFill(start_color="FFF6CC", end_color="FFF6CC", fill_type="solid")

    # v11 (2026-08-31, Andrew's ask): flags a fresh row whose player already has a
    # separate open "(reminder)" row elsewhere in the sheet, so the collision isn't easy to
    # miss. Kept short (no row number -- row numbers shift on every rewrite and would go
    # stale) since Andrew already knows to use the player-column filter to find the match.
    REMINDER_FLAG_PREFIX = "[Already has an open reminder] "
    REMINDER_FLAG_FONT = InlineFont(b=True)

    PENDING_SINCE_RE = re.compile(r"\s*\(pending since \d{4}-\d{2}-\d{2}\)\s*$")

    def _strip_pending_since(summary) -> str:
        """Old-format reminder rows (pre-2026-08-28) have '(pending since <date>)'
        baked into the summary text; new ones don't (the date column carries it
        instead). Stripped when comparing/migrating so an old-format row and a
        freshly-generated new-format summary for the same underlying story compare
        equal instead of looking like a spurious change."""
        return PENDING_SINCE_RE.sub("", str(summary or ""))

    # v17: fold in any rows queued by a previous run that couldn't save (see docstring)
    # -- each keeps the date it actually happened on, not today's date_str.
    queued_with_dates = []
    if VALIDATION_LOG_PENDING_PATH.exists():
        queued_with_dates = json.loads(VALIDATION_LOG_PENDING_PATH.read_text())
    all_with_dates = [(date_str, r) for r in changes] + [
        (item["date_str"], item["change"]) for item in queued_with_dates
    ]

    regular_rows = []
    for row_date, r in all_with_dates:
        headlines = " / ".join(it["headline"] for it in r["items"])
        regular_rows.append({
            "date": row_date, "player": r["full_name"], "position": r["position"],
            "team": r["team"] or "", "impact": r["impact"], "tone": r["tone"],
            "long_term": r.get("long_term", False), "summary": r["summary"],
            "headlines": headlines, "correct": "", "andrew_notes": "",
            "override_impact": "", "override_tone": "",
            "event_driven": r.get("event_driven", True),
            # TRIAL (2026-09-02): logged for Andrew's review only -- NOT yet used to
            # suppress a row. See parse_claude_response / append_validation_log docstring.
            "new_information": r.get("new_information", True),
        })

    # v11: previously bailed out here whenever today had nothing new, which also skipped
    # the auto-clear pass below on a quiet day -- a resolved reminder would sit stale until
    # the next day something DID change. Only bail early when there's truly nothing to do
    # (nothing new AND no existing file to clean up).
    if not regular_rows and not pending_reminders and not VALIDATION_LOG_PATH.exists():
        return 0

    VALIDATION_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    if VALIDATION_LOG_PATH.exists():
        wb = load_workbook(VALIDATION_LOG_PATH)
        ws = wb.active
        # Migrate an older workbook that predates a newly-added field (e.g.
        # event_driven, added 2026-08-27): append any VALIDATION_LOG_FIELDNAMES entry
        # missing from row 1 as a new header column, so existing files upgrade in
        # place instead of silently getting an unlabeled trailing column. Safe only
        # because new fields are always appended at the END of VALIDATION_LOG_FIELDNAMES
        # (Andrew's ask -- new columns go rightmost, existing layout stays put), so the
        # resulting header order always matches the field order used below.
        existing_headers = [c.value for c in ws[1]]
        for field in VALIDATION_LOG_FIELDNAMES:
            if field not in existing_headers:
                ws.cell(row=1, column=len(existing_headers) + 1, value=field)
                existing_headers.append(field)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(VALIDATION_LOG_FIELDNAMES)

    date_col = VALIDATION_LOG_FIELDNAMES.index("date") + 1
    impact_col = VALIDATION_LOG_FIELDNAMES.index("impact") + 1

    def _write_row(row_num: int, r: dict) -> None:
        for col, field in enumerate(VALIDATION_LOG_FIELDNAMES, start=1):
            ws.cell(row=row_num, column=col, value=r.get(field))
        date_val = r["date"]
        if isinstance(date_val, str):
            date_val = date.fromisoformat(date_val)
        elif isinstance(date_val, datetime):
            date_val = date_val.date()
        date_cell = ws.cell(row=row_num, column=date_col)
        date_cell.value = date_val
        # Write date as a real date type (not a string) so Excel can't silently
        # reformat it into something apply_validation_overrides.py can't parse.
        date_cell.number_format = "YYYY-MM-DD"
        if "reminder" in str(r.get("impact") or "").lower():
            for c in range(1, len(VALIDATION_LOG_FIELDNAMES) + 1):
                cell = ws.cell(row=row_num, column=c)
                cell.fill = REMINDER_FILL
                cell.border = REMINDER_BORDER

    # Read every existing data row back by VALUE (matched by header name, same approach
    # as apply_validation_overrides.py's load_rows) so it can be merged with the newly
    # generated rows and re-sorted -- see the v9 docstring note above on what this does
    # and does not preserve.
    existing_rows = []
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        existing_rows.append(dict(zip(headers, row)))

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # v10 one-time migration: collapse any player with MULTIPLE existing "(reminder)"
    # rows (left over from the old once-a-day-forever behavior) down to one, keeping
    # whichever duplicate carries the most of Andrew's own content and refreshing its
    # summary/date to the most recent of the group (existing_rows preserves the sheet's
    # prior chronological order, so the group's last item is the most recent).
    # v11: a player's CURRENT true status per history.json -- used below to auto-clear a
    # reminder row once the underlying story has actually resolved (Andrew's ask, Option A:
    # he accepted that a resolved row's note/override is gone with it, not archived).
    name_to_latest_impact = {
        rec["full_name"]: rec["entries"][-1]["impact"]
        for rec in history.values() if rec.get("entries")
    }

    other_existing = []
    reminder_groups: dict = {}
    for r in existing_rows:
        if "reminder" in str(r.get("impact") or "").lower():
            reminder_groups.setdefault(r.get("player"), []).append(r)
        else:
            other_existing.append(r)

    def _annotation_score(r: dict) -> int:
        return sum(1 for f in ("correct", "andrew_notes", "override_impact", "override_tone") if r.get(f))

    open_reminders_by_player: dict = {}
    for player, group in reminder_groups.items():
        # v11 auto-clear: the story has resolved (player's latest recorded status isn't
        # PENDING any more, or they've dropped out of history entirely) -- drop the row
        # rather than carry a stale reminder forward.
        if name_to_latest_impact.get(player) != "PENDING":
            continue
        keeper = max(group, key=_annotation_score) if len(group) > 1 else group[0]
        latest = group[-1]
        keeper["summary"] = _strip_pending_since(latest["summary"])
        keeper["date"] = latest["date"]
        open_reminders_by_player[player] = keeper

    # v15 (2026-09-02, bug found via a real Malik Nabers row): snapshot the set of
    # players with a SEPARATE, PRE-EXISTING open reminder BEFORE the upsert loop below
    # runs. The collision flag below must only fire against this snapshot, not the
    # post-upsert dict -- a player going PENDING for the very first time gets a regular
    # row (in `changes`) AND a brand-new reminder row in this SAME call, and checking
    # post-upsert made that brand-new reminder collide with the row that created it,
    # flagging "[Already has an open reminder]" on a story that in fact has no separate
    # reminder anywhere -- it just became one this run. Caught 2026-09-01 in real data:
    # Nabers' first PENDING classification got flagged against itself; Andrew had
    # already marked that row incorrect (correct='N') before this was root-caused.
    pre_existing_reminder_players = set(open_reminders_by_player.keys())

    # v10 upsert: a still-open pending story updates its ONE existing reminder row in
    # place (summary/date refreshed, Andrew's correct/andrew_notes/override_* left
    # exactly as they were) instead of adding a new row every day nothing has changed.
    # Only a player with no existing open reminder row gets a brand new one.
    new_reminder_count = 0
    for p in pending_reminders:
        existing = open_reminders_by_player.get(p["full_name"])
        if existing is not None:
            existing["summary"] = p["summary"]
            existing["date"] = p["since_date"]
        else:
            open_reminders_by_player[p["full_name"]] = {
                "date": p["since_date"], "player": p["full_name"], "position": p["position"],
                "team": p["team"] or "", "impact": "PENDING (reminder)", "tone": "",
                "long_term": "", "summary": p["summary"], "headlines": "", "correct": "",
                "andrew_notes": "", "override_impact": "", "override_tone": "", "event_driven": "",
            }
            new_reminder_count += 1

    # v11: flag a fresh regular row when its player ALSO still has a SEPARATE,
    # PRE-EXISTING open reminder row (post auto-clear above, so a just-resolved story
    # never gets flagged against itself; pre-upsert snapshot, so a story becoming
    # PENDING for the first time never gets flagged against the reminder it just
    # created for itself -- see v15 note above) -- bolded prefix via rich text, see
    # REMINDER_FLAG_PREFIX above.
    for r in regular_rows:
        if r["player"] in pre_existing_reminder_players:
            r["summary"] = CellRichText(
                TextBlock(REMINDER_FLAG_FONT, REMINDER_FLAG_PREFIX), str(r["summary"] or "")
            )

    # Stable sort: every non-reminder row (existing rows in their original order, then
    # this run's new ones) before every reminder row (same) -- Andrew's ask, 2026-08-27,
    # so reminders always end up grouped at the true bottom of the whole sheet.
    merged = other_existing + regular_rows + list(open_reminders_by_player.values())
    merged.sort(key=lambda r: "reminder" in str(r.get("impact") or "").lower())

    for i, r in enumerate(merged, start=2):
        _write_row(i, r)

    for i, field in enumerate(VALIDATION_LOG_FIELDNAMES, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = VALIDATION_LOG_COLUMN_WIDTHS[field]
    ws.freeze_panes = "A2"  # keep the header row visible while scrolling -- Andrew's ask, 2026-08-25

    # AutoFilter dropdowns on every column (Andrew's ask, 2026-08-26) -- lets him filter
    # `impact` down to just "PENDING"/"PENDING (reminder)" to see what still needs a
    # verdict, filter `player` to jump straight to one name, or filter to a specific
    # UP/DOWN/NEUTRAL to spot-check a category, instead of scrolling the whole sheet.
    # Re-applied on every write (like the column widths/freeze pane above) so it always
    # covers the full current row range, not just whatever it was when the file was
    # first created.
    last_col_letter = get_column_letter(len(VALIDATION_LOG_FIELDNAMES))
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    # "Needs review" highlight (Andrew's ask, 2026-08-27): a live conditional-formatting
    # rule rather than a one-time fill, so it tracks his edits automatically -- a
    # non-reminder row highlights while `correct` is still blank and clears the moment
    # he fills it in, with no extra script run required. Reminder rows are excluded
    # (their `correct` isn't meant to be filled in the normal case) and already carry
    # their own grey fill. Cleared and re-added on every write so the range always
    # covers the full current sheet, not just whatever it was when this rule was first
    # added.
    correct_letter = get_column_letter(VALIDATION_LOG_FIELDNAMES.index("correct") + 1)
    impact_letter = get_column_letter(impact_col)
    ws.conditional_formatting = ConditionalFormattingList()
    ws.conditional_formatting.add(
        f"A2:{last_col_letter}{ws.max_row}",
        FormulaRule(
            formula=[f'AND(${correct_letter}2="",ISERROR(SEARCH("reminder",${impact_letter}2)))'],
            fill=NEEDS_REVIEW_FILL,
        ),
    )

    # Override guardrail (Andrew's ask, 2026-08-27): the grey fill alone hasn't stopped
    # an override from landing on a "(reminder)" row -- happened more than once (see
    # apply_validation_overrides.py's docstring, and Malik Nabers 2026-08-27). This
    # rejects it at entry time instead of relying on Andrew remembering the rule: Excel
    # data validation on override_impact/override_tone that hard-stops any value typed
    # into those cells on a reminder row. Cleared and re-added on every write so the
    # range covers the full current sheet.
    ws.data_validations = DataValidationList()
    override_impact_letter = get_column_letter(VALIDATION_LOG_FIELDNAMES.index("override_impact") + 1)
    override_tone_letter = get_column_letter(VALIDATION_LOG_FIELDNAMES.index("override_tone") + 1)
    dv = DataValidation(
        type="custom",
        formula1=f'ISERROR(SEARCH("reminder",${impact_letter}2))',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Wrong row for an override",
        error=('Overrides do not apply on "(reminder)" rows -- apply_validation_overrides.py '
               "skips them by design. Put this on the player's original entry instead (the "
               "plain-value row where this status was first classified)."),
    )
    dv.add(f"{override_impact_letter}2:{override_tone_letter}{ws.max_row}")
    ws.add_data_validation(dv)

    # v17: a couple of quick retries covers a brief transient lock; the common real
    # case (Andrew has the file open in Excel) won't clear in seconds, so this isn't
    # meant to wait him out -- it's the queue-on-failure below that actually handles that.
    for attempt in range(3):
        try:
            wb.save(VALIDATION_LOG_PATH)
            break
        except PermissionError:
            if attempt == 2:
                VALIDATION_LOG_PENDING_PATH.write_text(json.dumps(
                    [{"date_str": row_date, "change": r} for row_date, r in all_with_dates]
                ))
                print(
                    f"WARNING: {VALIDATION_LOG_PATH.name} is open elsewhere (likely in "
                    f"Excel) -- {len(all_with_dates)} row(s) queued to "
                    f"{VALIDATION_LOG_PENDING_PATH.name} and will be written automatically "
                    "on the next run that can save."
                )
                return 0
            time.sleep(3)
    if VALIDATION_LOG_PENDING_PATH.exists():
        VALIDATION_LOG_PENDING_PATH.unlink()
    # Counts brand-new rows only -- a pending story that just refreshed an existing
    # reminder row in place (v10) isn't a "new" entry the same way a fresh regular row
    # or a first-time reminder is.
    return len(regular_rows) + new_reminder_count


def main() -> None:
    load_dotenv()
    today = datetime.now(timezone.utc).date()
    today_str = today.isoformat()

    conn = sqlite3.connect(DB_PATH)
    try:
        name_lookup = build_name_lookup(conn)
    finally:
        conn.close()

    history = prune_history(load_history(), today)

    print("Fetching RotoWire NFL news feed...")
    rotowire_items = fetch_rotowire_items()
    print(f"  {len(rotowire_items)} items in feed")

    print("Fetching RotoBaller player-news feed...")
    rotoballer_items = fetch_rotoballer_items()
    print(f"  {len(rotoballer_items)} football-tagged items in feed")

    print("Matching items to known players...")
    matched = match_rotowire_items(rotowire_items, name_lookup)
    matched += match_rotoballer_items(rotoballer_items, name_lookup)

    if not matched:
        print("No items matched a known player -- nothing to summarize.")
        save_history(history)  # still persist any pruning that happened above
        return

    groups = group_by_player(matched)
    merged_count = len(matched) - len(groups)
    if merged_count:
        print(f"  Merged {merged_count} duplicate item(s) reporting the same player from multiple sources")

    client = Anthropic()  # reads ANTHROPIC_API_KEY from the environment (loaded above via .env)
    for group in groups:
        sources = ", ".join(it["source"] for it in group["items"])
        headlines = " / ".join(it["headline"] for it in group["items"])
        team_tag = group["team"] or "FA"
        print(f"  Summarizing ({sources}): {group['full_name']} ({team_tag}-{group['position']}) -- {headlines}")
    # v14: one batched call for the whole run instead of one call per player (see
    # summarize_batch_with_claude docstring) -- falls back to individual calls per
    # player internally if anything about the batch response can't be trusted.
    results = summarize_batch_with_claude(client, groups, history)

    changes = [r for r in results if is_change(r, history)]
    print(f"  {len(changes)} of {len(results)} are new/changed since last known status")

    history = update_history(history, results, today_str)
    pending_reminders = get_pending_reminders(history, today_str)
    save_history(history)

    out_path = write_report(changes, pending_reminders, today_str)
    if out_path:
        print(f"\nWrote report to {out_path}")
    else:
        print("\nNothing new to report (no changes, no unreminded pending items) -- no file written.")

    logged_count = append_validation_log(changes, pending_reminders, today_str, history)
    if logged_count:
        print(f"Logged {logged_count} entries to {VALIDATION_LOG_PATH} for review "
              f"(fill in the 'correct' and 'andrew_notes' columns after reading the report)")


if __name__ == "__main__":
    main()
