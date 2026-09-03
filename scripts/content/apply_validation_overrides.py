"""
Applies Andrew's corrections from digest_validation_log.xlsx into digest_history.json.

Why this exists (2026-08-25): the digest's IMPACT/TONE classification is written by
Claude and immediately becomes "prior status" that future runs trust (see digest.py's
format_history_for_prompt). If the AI's call was wrong and never corrected, that wrong
call just propagates forward untouched. Andrew's ask: he wants final say on
impact/tone, not just a passive log of what the AI decided.

How to use: after reading a digest report, fill in `override_impact` and/or
`override_tone` in data/processed/digest_validation_log.xlsx for any row where the AI's
call was wrong (leave blank to accept the AI's call as-is). Then run this script. It
finds the matching entry in digest_history.json (matched by player full_name + date +
summary) and overwrites impact/tone with your correction, so the NEXT digest run sees
the corrected status, not the AI's original one.

Safe to run repeatedly -- re-applying the same override is a no-op, and rows without
an override are skipped entirely. Doesn't touch the workbook itself, only history.json,
so there's no "did this get applied already" bookkeeping to worry about.

v2 (2026-08-25, same-day format switch): originally read a CSV, but Excel silently
reformatted the date column on save (2026-08-25 -> 8/25/2026), which broke date
matching entirely -- every row SKIPped on Andrew's first real run. digest.py now writes
this file as .xlsx with the date column as a real Excel date type instead of a string,
which openpyxl reads back as an actual datetime object regardless of how Excel chose to
DISPLAY it -- more robust than string-parsing multiple date formats. Kept a string
fallback below in case a date cell ever ends up as plain text (e.g. typed by hand).

Matches this project's rule against writing to data/db/*.db or similar file-locking-
sensitive paths from the Claude sandbox -- run this locally, not from Claude's sandbox.

Usage:
    py scripts/content/apply_validation_overrides.py
"""

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
HISTORY_PATH = PROJECT_ROOT / "data" / "processed" / "digest_history.json"
VALIDATION_LOG_PATH = PROJECT_ROOT / "data" / "processed" / "digest_validation_log.xlsx"

VALIDATION_LOG_FIELDNAMES = [
    "date", "player", "position", "team", "impact", "tone", "long_term",
    "summary", "headlines", "correct", "andrew_notes", "override_impact", "override_tone",
    "event_driven",
]

# String fallback only -- normally the date column is a real Excel date type (see
# module docstring), but this covers a cell that ended up as plain text.
DATE_INPUT_FORMATS = ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"]


def normalize_date(raw) -> str | None:
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    raw = str(raw or "").strip()
    for fmt in DATE_INPUT_FORMATS:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def load_rows() -> list[dict]:
    """Reads every data row out of the workbook as a plain dict keyed by
    VALIDATION_LOG_FIELDNAMES, regardless of column order in the header row (matches
    by header name, not position, so this survives Andrew reordering columns)."""
    if not VALIDATION_LOG_PATH.exists():
        print(f"No validation log found at {VALIDATION_LOG_PATH} -- nothing to apply.")
        return []
    wb = load_workbook(VALIDATION_LOG_PATH, data_only=True)
    ws = wb.active

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(header, row)))
    return rows


def load_overrides() -> list[dict]:
    """Only rows with a non-blank override_impact or override_tone matter here --
    everything else in the workbook (correct, andrew_notes on rows with no override) is
    Andrew's own record-keeping, not something this script acts on.

    "PENDING (reminder)" rows are explicitly skipped -- they're not a separate AI
    classification, just a re-surfaced notice about an ALREADY-existing entry (see
    digest.py's get_pending_reminders). Overriding one doesn't correct anything real; it
    risks colliding with -- and silently overwriting -- an override already given (or
    correctly left blank) on that player's actual entry for the same date. Found this
    2026-08-25 reviewing Andrew's filled-in log: TreVeyon Henderson's real row was
    marked correct (no override needed), but the reminder row for the same player/date
    had override_impact=Down filled in -- applying both would have let the reminder
    row's override silently clobber the real row's "leave it alone" verdict."""
    rows = load_rows()

    skipped_reminders = 0
    overrides = []
    for r in rows:
        has_override = str(r.get("override_impact") or "").strip() or str(r.get("override_tone") or "").strip()
        if not has_override:
            continue
        if "reminder" in str(r.get("impact") or "").lower():
            skipped_reminders += 1
            continue
        overrides.append(r)

    if skipped_reminders:
        print(f"Skipping {skipped_reminders} override(s) on 'PENDING (reminder)' rows -- "
              f"these reference an already-existing entry, not a new classification to "
              f"correct. If you meant to correct that player's actual status, put the "
              f"override on their real (non-reminder) row for that date instead.")
    return overrides


def apply_overrides(history: dict, overrides: list[dict]) -> int:
    applied = 0
    for row in overrides:
        player = row["player"]
        summary = row.get("summary", "")
        date_str = normalize_date(row.get("date"))
        if date_str is None:
            print(f"  SKIP: {player} ({row.get('date')!r}) -- couldn't parse this date, "
                  f"check the date column wasn't accidentally typed as text")
            continue

        # Match by full_name -- the workbook doesn't carry gsis_id. Same pragmatic
        # "duplicate names are rare enough to accept" tradeoff already made elsewhere
        # in this project (see ffc.py's normalize_name matching).
        record = next((r for r in history.values() if r["full_name"] == player), None)
        if record is None:
            print(f"  SKIP: {player} ({date_str}) -- not found in history.json")
            continue

        matching_entries = [e for e in record["entries"] if e["date"] == date_str]
        if not matching_entries:
            print(f"  SKIP: {player} ({date_str}) -- no history entry for that date")
            continue

        # A player can have MULTIPLE entries on the same calendar date if the digest
        # ran more than once that day (e.g. Ashton Jeanty had both an UP entry and a
        # later NEUTRAL "on the mend" entry, both dated 2026-08-25). Matching on date
        # alone would silently apply the override to whichever entry happens to be
        # LATEST, which may not be the one the row was actually about. Match on the
        # exact summary text first to disambiguate; only fall back to "last entry for
        # the date" -- with a visible warning -- if that fails (e.g. summary got edited
        # in the spreadsheet).
        exact = [e for e in matching_entries if e.get("summary") == summary]
        if exact:
            entry = exact[-1]
        else:
            entry = matching_entries[-1]
            if len(matching_entries) > 1:
                print(f"  WARNING: {player} ({date_str}) has {len(matching_entries)} entries that "
                      f"date and none matched this row's summary exactly -- applying to the most "
                      f"recent one. Double-check this was the right entry.")

        new_impact = str(row.get("override_impact") or "").strip().upper()
        new_tone = str(row.get("override_tone") or "").strip().upper()
        changed = []
        if new_impact and entry["impact"] != new_impact:
            changed.append(f"impact {entry['impact']} -> {new_impact}")
            entry["impact"] = new_impact
        if new_tone and entry["tone"] != new_tone:
            changed.append(f"tone {entry['tone']} -> {new_tone}")
            entry["tone"] = new_tone

        if changed:
            print(f"  {player} ({date_str}): {', '.join(changed)}")
            applied += 1
        else:
            print(f"  {player} ({date_str}): override matches recorded value already, no change")

    return applied


def main() -> None:
    overrides = load_overrides()
    if not overrides:
        print("No override_impact/override_tone values filled in -- nothing to do.")
        return

    print(f"Found {len(overrides)} row(s) with an override. Applying...")
    history = json.loads(HISTORY_PATH.read_text())
    applied = apply_overrides(history, overrides)

    if applied:
        HISTORY_PATH.write_text(json.dumps(history, indent=2))
        print(f"\nUpdated {applied} entr{'y' if applied == 1 else 'ies'} in {HISTORY_PATH}")
    else:
        print("\nNo changes needed -- history.json already matches your overrides.")


if __name__ == "__main__":
    main()
