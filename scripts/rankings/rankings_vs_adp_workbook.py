"""
Builds a polished Excel workbook comparing our rankings against market ADP -- a more
readable evolution of export_rankings_vs_adp.py's CSV (Andrew's ask, 2026-08-02):
color-graded by how far off the market we are, instead of a flat table of numbers.

Overall tab: Rank, Position, Name, ADP, Delta -- cross-position comparison, sorted by
our overall rank.

Position tabs (QB/RB/WR/TE -- K excluded, same reasoning as export_rankings_vs_adp.py:
kicker ADP is thin/unreliable and not where the interesting disagreements live): Our Pos
Rank (RB1, RB2, ...) vs ADP Pos Rank (the market's OWN positional ordering, e.g. "the
37th cheapest RB by ADP" -- not the same number as the Overall tab's raw-ADP-value delta).

v2 (2026-08-24, Andrew's ask): each position tab now shows the UNION of our own top N
(POSITION_LIMITS, unchanged) and EVERY player the market drafts at that position (ADP
Pos Rank is now computed over the market's full pool, not just players we also rank --
the whole point is to surface players we're missing, so restricting the market side to
"players we already rank" would have hidden exactly the players this feature exists to
find). A player with ADP who isn't in our top N shows a Status explaining why:
  - "Below our top N" -- we DO score them, just deeper than our display cutoff. An
    extend-our-depth candidate, not a real gap.
  - "No current team" -- excluded by scoring.py's roster filter. Likely a stale FFC
    mock-draft entry (same class of issue the Njoku/Waller/Ertz cases surfaced earlier),
    but worth a glance in case our own roster data (sleeper.py) is what's stale instead.
  - "Rookie, no draft capital (UDFA)" -- this year's class, but scoring.py's rookie path
    requires a real draft pick to build a baseline off of (see get_rookie_candidates).
  - "No games since <season> (2+ seasons inactive)" -- excluded by the recency
    eligibility filter; a player the market still mock-drafts out of habit/name value.
  - "Not scored -- reason unclear, worth checking" -- none of the above applied, which
    shouldn't normally happen; a genuine "go look at the data" flag rather than a guess.
These mirror scoring.py's actual filter order (current-roster check, then rookie/draft-
capital check, then recency eligibility) so the explanation is read off the real
pipeline logic, not a separate guess that could drift out of sync with it.

v3 (2026-08-24, Andrew's ask): the Overall tab gets the same treatment -- Rank shows
"NR" for a market-drafted player we never scored at all (see write_overall_tab). Kept
narrower than the position tabs, though: a player we DO score but who falls outside our
own top-N display window still isn't added here, since that's a different question
("should our depth go deeper") than what was actually asked ("show me who I don't have
ranked").

Delta sign convention matches adp_comparison.py throughout: positive = we rank the
player BETTER than the market does (possible value); negative = we rank them WORSE
(possible fade or blind spot).

Color scale: a real 3-color Excel ColorScaleRule (NOT a FormulaRule -- see
cheat_sheet.py's note on why FormulaRule fills need both fgColor/bgColor set, a real bug
caught earlier this project. ColorScaleRule is a different XML structure entirely and
doesn't have that gotcha -- confirmed by round-tripping the saved file and inspecting the
actual color stops before shipping this, not just assumed). Dark red at the most-negative
delta in that column, white at exactly 0, dark green at the most-positive delta. Each tab
scales independently against its own min/max, so contrast is always meaningful for that view.

Usage:
    py scripts/rankings/rankings_vs_adp_workbook.py
"""

import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from cheat_sheet import fetch_rankings, DB_PATH, PROJECT_ROOT
from blurb_worklist import fetch_adp_lookup
from scoring import REFERENCE_SEASON
from weights_config import WEIGHTS

OUTPUT_PATH = PROJECT_ROOT / "content" / "rankings_vs_adp.xlsx"

POSITION_LIMITS = {
    "QB": 20,
    "RB": 40,
    "WR": 45,
    "TE": 20,
}

HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="404040")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
DECIMAL_FORMAT = "0.0"

# Excel's own default red/white/green 3-color scale hex values -- native, familiar look.
COLOR_RED = "F8696B"
COLOR_WHITE = "FFFFFF"
COLOR_GREEN = "63BE7B"


def apply_delta_color_scale(ws, delta_col_idx: int, first_row: int, last_row: int) -> None:
    """3-color scale on the Delta column: dark red at this column's most-negative value,
    white fixed at exactly 0 (not the range midpoint -- 0 is the meaningful "market and
    us agree" point), dark green at the most-positive value."""
    col_letter = get_column_letter(delta_col_idx)
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="min", start_color=COLOR_RED,
            mid_type="num", mid_value=0, mid_color=COLOR_WHITE,
            end_type="max", end_color=COLOR_GREEN,
        ),
    )


def fetch_adp_players_by_position(conn: sqlite3.Connection, position: str) -> dict:
    """Every player at this position with a current FFC ADP -- the market's FULL
    mock-draft pool for the position, regardless of whether we've scored them. This is
    deliberately NOT restricted to players we also rank (v1 of this script was, to avoid
    stale-ADP noise -- see module docstring) -- v2's whole point is to surface players
    we're missing, so the market side needs to be uncapped. Returns {gsis_id: dict},
    each tagged with adp_pos_rank (1-indexed, ascending ADP = better market rank)."""
    cur = conn.execute(
        """
        SELECT a.gsis_id, p.full_name, p.current_team, p.draft_year, p.draft_pick, a.adp
        FROM adp a
        JOIN players p ON a.gsis_id = p.gsis_id
        WHERE a.source = 'ffc' AND p.position = ?
          AND a.pulled_at = (SELECT MAX(pulled_at) FROM adp WHERE source = 'ffc')
        ORDER BY a.adp ASC
        """,
        (position,),
    )
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, row)) for row in cur.fetchall()]
    for i, r in enumerate(rows, start=1):
        r["adp_pos_rank"] = i
    return {r["gsis_id"]: r for r in rows}


def fetch_most_recent_season_by_player(conn: sqlite3.Connection) -> dict:
    """gsis_id -> last season with a logged game -- used to explain the recency-
    eligibility exclusion the same way scoring.py's filter_eligible_players checks it."""
    cur = conn.execute("SELECT gsis_id, MAX(season) FROM weekly_stats GROUP BY gsis_id")
    return dict(cur.fetchall())


def determine_missing_reason(adp_player: dict, most_recent_season_by_player: dict) -> str:
    """Why a market-drafted player isn't in our rankings -- checked in the same order
    scoring.py's own filters run (current-roster, then rookie/draft-capital, then
    recency eligibility), so this stays consistent with the real pipeline logic instead
    of being a separate guess that could quietly drift out of sync with it."""
    if not adp_player["current_team"]:
        return "No current team (free agent/unsigned)"
    if adp_player["draft_year"] == REFERENCE_SEASON:
        return ("Rookie, no draft capital (UDFA)" if adp_player["draft_pick"] is None
                else "Rookie -- check data (has draft capital, should be scored)")
    most_recent = most_recent_season_by_player.get(adp_player["gsis_id"])
    cutoff = REFERENCE_SEASON - WEIGHTS["eligibility_window_seasons"]
    if most_recent is None:
        return "No games in our data"
    if most_recent < cutoff:
        return f"No games since {most_recent} (2+ seasons inactive)"
    return "Not scored -- reason unclear, worth checking"


def write_overall_tab(ws, all_rankings: list[dict], adp_lookup: dict, adp_by_position: dict,
                       most_recent_season_by_player: dict) -> None:
    """Rank shows "NR" (not ranked) for a player the market drafts that we don't score
    at all -- Andrew's ask, 2026-08-24, same gap as the position tabs had before v2: this
    tab previously only ever showed OUR top-N-per-position, so an unscored player (free
    agent, UDFA, inactive veteran -- see determine_missing_reason) could never appear
    here no matter how high the market has them. Deliberately narrower than the position
    tabs' "Below our top N" case, though: a player we DO score but who falls outside our
    top-N display window still isn't shown here at all -- this only adds players we
    never ranked, full stop, since that's specifically what was asked for."""
    columns = [
        ("Rank", 8), ("Position", 10), ("Name", 25), ("Team", 8), ("ADP", 9), ("Delta", 9), ("Status", 38),
    ]
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    display_pool = []
    for position, limit in POSITION_LIMITS.items():
        position_rows = sorted(
            (r for r in all_rankings if r["position"] == position),
            key=lambda r: r["positional_rank"],
        )[:limit]
        display_pool.extend(position_rows)

    rows = []
    for r in display_pool:
        adp = adp_lookup.get(r["gsis_id"])
        rows.append({
            "rank": r["rank"], "position": r["position"], "name": r["full_name"], "team": r["current_team"],
            "adp": round(adp, 1) if adp is not None else None,
            "delta": round(adp - r["rank"], 1) if adp is not None else None,
            "status": None,
            "sort_key": adp if adp is not None else r["rank"] + 100000,  # unmatched-ADP rows sort after everything ADP'd
        })

    our_scored_ids = {r["gsis_id"] for r in all_rankings}
    for position, adp_players in adp_by_position.items():
        for gsis_id, adp_p in adp_players.items():
            if gsis_id in our_scored_ids:
                continue  # already covered above -- either displayed or below our own cutoff, not "unranked"
            rows.append({
                "rank": "NR", "position": position, "name": adp_p["full_name"], "team": adp_p["current_team"],
                "adp": round(adp_p["adp"], 1), "delta": None,
                "status": determine_missing_reason(adp_p, most_recent_season_by_player),
                "sort_key": adp_p["adp"],
            })

    rows.sort(key=lambda r: r["sort_key"])

    row_idx = 2
    for r in rows:
        values = [r["rank"], r["position"], r["name"], r["team"], r["adp"], r["delta"], r["status"]]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BOLD_FONT if col_idx == 3 else BODY_FONT
            if col_idx in (5, 6):
                cell.number_format = DECIMAL_FORMAT
        row_idx += 1

    last_row = row_idx - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_row}"
    apply_delta_color_scale(ws, delta_col_idx=6, first_row=2, last_row=last_row)


def write_position_tab(ws, all_rankings: list[dict], adp_by_position: dict,
                        most_recent_season_by_player: dict, position: str, limit: int) -> None:
    columns = [
        ("Our Pos Rank", 13), ("ADP Pos Rank", 13), ("Name", 25), ("Team", 8),
        ("ADP", 9), ("Delta", 9), ("Status", 38),
    ]
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # our_players: everyone we've scored at this position (any depth) -- needed so a
    # below-cutoff player still gets their real "Our Pos Rank" and correct Status,
    # rather than being treated as unscored. our_top_n: just the display-limit subset,
    # used only to decide which of our own players pull a row into the union below.
    our_players = {r["gsis_id"]: r for r in all_rankings if r["position"] == position}
    our_top_n_ids = {gsis_id for gsis_id, r in our_players.items() if r["positional_rank"] <= limit}
    adp_players = adp_by_position  # already scoped to this position by the caller

    display_rows = []
    for gsis_id in our_top_n_ids | set(adp_players):
        our = our_players.get(gsis_id)
        adp_p = adp_players.get(gsis_id)

        our_label = f"{position}{our['positional_rank']}" if our else None
        adp_label = f"{position}{adp_p['adp_pos_rank']}" if adp_p else None
        adp_value = adp_p["adp"] if adp_p else None
        name = our["full_name"] if our else adp_p["full_name"]
        team = our["current_team"] if our else adp_p["current_team"]

        if our and adp_p:
            delta = adp_p["adp_pos_rank"] - our["positional_rank"]
            status = None if our["positional_rank"] <= limit else f"Below our top {limit}"
        elif our:  # our top-N player, no market ADP coverage at all -- not an issue, just unmocked
            delta = None
            status = None
        else:  # market drafts them, we don't score them at all -- the interesting case
            delta = None
            status = determine_missing_reason(adp_p, most_recent_season_by_player)

        display_rows.append({
            "our_label": our_label, "adp_label": adp_label, "name": name, "team": team,
            "adp": round(adp_value, 1) if adp_value is not None else None, "delta": delta, "status": status,
            "sort_key": adp_value if adp_value is not None else 9999,
        })

    display_rows.sort(key=lambda r: r["sort_key"])

    row_idx = 2
    for r in display_rows:
        values = [r["our_label"], r["adp_label"], r["name"], r["team"], r["adp"], r["delta"], r["status"]]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BOLD_FONT if col_idx == 3 else BODY_FONT
            if col_idx == 5:
                cell.number_format = DECIMAL_FORMAT
        row_idx += 1

    last_row = row_idx - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_row}"
    apply_delta_color_scale(ws, delta_col_idx=6, first_row=2, last_row=last_row)


def main() -> None:
    conn = sqlite3.connect(DB_PATH)
    try:
        all_rankings = fetch_rankings(conn)
        adp_lookup = fetch_adp_lookup(conn)
        most_recent_season_by_player = fetch_most_recent_season_by_player(conn)
        adp_by_position = {
            position: fetch_adp_players_by_position(conn, position)
            for position in POSITION_LIMITS
        }
    finally:
        conn.close()

    if not all_rankings:
        raise SystemExit("No rankings found -- run scoring.py first.")

    wb = Workbook()
    overall_ws = wb.active
    overall_ws.title = "Overall"
    write_overall_tab(overall_ws, all_rankings, adp_lookup, adp_by_position, most_recent_season_by_player)

    missing_counts = {}
    for position, limit in POSITION_LIMITS.items():
        ws = wb.create_sheet(position)
        write_position_tab(ws, all_rankings, adp_by_position[position],
                            most_recent_season_by_player, position, limit)
        our_ids = {r["gsis_id"] for r in all_rankings if r["position"] == position}
        missing_counts[position] = len(set(adp_by_position[position]) - our_ids)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote Overall + {len(POSITION_LIMITS)} position tabs to {OUTPUT_PATH}")
    print("Market-drafted players we don't score at all, by position:", missing_counts)


if __name__ == "__main__":
    main()
