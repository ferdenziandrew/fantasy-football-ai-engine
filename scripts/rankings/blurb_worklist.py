"""
Builds a lightweight, editable worksheet for writing player blurbs (the "why" behind a
ranking) -- NOT the final cheat sheet. Surfaces the signals worth writing about (ADP
disagreement, rookie flag, recent missed games) right next to each name, so writing a
blurb doesn't require digging back through the cheat sheet or DB while doing it.

Workflow (ROADMAP.md Phase 2 / CLAUDE.md): Andrew fills in Notes for players he has a
real opinion on and leaves the rest blank; Claude drafts the blanks from these same
signals; Andrew reviews/approves before anything loads into rankings.blurb/blurb_source.
This script only produces the worksheet -- loading approved blurbs back into the DB
(matching rows by the hidden gsis_id column, not by name -- see ROADMAP.md's
"Mike Washington Jr." duplicate-name note for why name-matching isn't safe) is a
separate step once the worksheet comes back filled in.

Reuses cheat_sheet.py's fetch_rankings()/compute_score_gap_tiers() rather than
re-deriving rank/tier/rookie logic, so this worklist and the cheat sheet can never
quietly disagree about a player's tier or rank. ADP/delta join mirrors
adp_comparison.py's methodology (FFC source, latest pull, delta = adp - our_rank) for
the same reason.

"Missed Gm '25" is a simple, transparent injury-history signal: 17 (standard 2021+
regular season length) minus the player's 2025 games_played, floored at 0. Not shown
for rookies (they have zero 2025 games by definition -- "missed 17 games" would read as
an injury flag when it's really just "hadn't been drafted yet").

Usage:
    py scripts/rankings/blurb_worklist.py
"""

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from cheat_sheet import fetch_rankings, compute_score_gap_tiers, DB_PATH, PROJECT_ROOT

# Each position pass gets its own output file rather than one shared file that gets
# regenerated -- QB/TE blurbs are already written (content/blurb_worklist_v2.xlsx,
# Andrew's edits + typo fixes applied 2026-07-30) but NOT YET loaded back into
# rankings.blurb, so re-running this script with QB/TE still in POSITION_LIMITS would
# have overwritten that file with blank Notes again. RB/WR is a new, separate pass;
# once all four positions are loaded back into the DB (see the not-yet-built
# "load worklist back into rankings.blurb" step), a single combined file becomes
# possible again if wanted.
OUTPUT_PATH = PROJECT_ROOT / "content" / "blurb_worklist_rbwr.xlsx"

SCORING_FORMAT = "ppr"
ADP_SOURCE = "ffc"
FULL_SEASON_GAMES = 17  # 2021+ NFL regular season length -- basis for the missed-games signal

# Position -> how many of that position's top players (by positional_rank) to include
# in this pass. Depth varies by position on purpose -- RB/WR run much deeper in real
# 10-12 team draft relevance (flex spots, handcuffs, deep WR3/4 in PPR) than QB/TE did.
# Starting point, adjust after seeing how it feels to write against: RB to 40 (flex/
# handcuff relevant), WR to 50 (deepest position in PPR).
POSITION_LIMITS = {
    "RB": 40,
    "WR": 50,
}

HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
DECIMAL_FORMAT = "0.0"

# (header, dict key, width, is_decimal) -- gsis_id is last and hidden: not for Andrew to
# edit, just so a later "load blurbs back into the DB" step can match rows exactly
# instead of by name (see the duplicate-name gotcha noted in ROADMAP.md).
COLUMNS = [
    ("Rank", "rank", 8, False),
    ("Player", "full_name", 25, False),
    ("Team", "current_team", 8, False),
    ("Pos Rank", "pos_rank_label", 11, False),
    ("Tier", "tier_num", 8, False),
    ("ADP", "adp", 9, True),
    ("Delta", "delta", 9, True),
    ("Rookie", "rookie_label", 9, False),
    ("Missed Gm '25", "missed_games_label", 14, False),
    ("Notes", "blurb", 60, False),
    ("gsis_id", "gsis_id", 14, False),
]


def fetch_adp_lookup(conn: sqlite3.Connection) -> dict:
    """gsis_id -> latest FFC adp value. Same source/recency rule as adp_comparison.py,
    so the two tools never quietly disagree about what "the market" says."""
    cur = conn.execute(
        """
        SELECT gsis_id, adp FROM adp
        WHERE source = ? AND pulled_at = (SELECT MAX(pulled_at) FROM adp WHERE source = ?)
        """,
        (ADP_SOURCE, ADP_SOURCE),
    )
    return dict(cur.fetchall())


def build_position_rows(all_rows: list[dict], adp_lookup: dict, position: str, limit: int) -> list[dict]:
    position_rows = [r for r in all_rows if r["position"] == position]

    # Tiers must be computed over the FULL position (gaps look different over 20 players
    # vs. the whole position pool), then trimmed to the top N after -- not the other way
    # around, or the tier boundaries here would silently disagree with the cheat sheet's.
    full_sorted_by_score = sorted(position_rows, key=lambda r: r["score"], reverse=True)
    tier_indices = compute_score_gap_tiers(full_sorted_by_score)
    for r, tier_idx in zip(full_sorted_by_score, tier_indices):
        r["tier_num"] = tier_idx + 1

    top_n = sorted(position_rows, key=lambda r: r["positional_rank"])[:limit]

    for r in top_n:
        adp = adp_lookup.get(r["gsis_id"])
        r["adp"] = round(adp, 1) if adp is not None else None
        r["delta"] = round(adp - r["rank"], 1) if adp is not None else None
        r["rookie_label"] = "Yes" if r["is_rookie_baseline"] else ""
        if r["is_rookie_baseline"]:
            r["missed_games_label"] = "-- (rookie)"
        else:
            missed = max(0, FULL_SEASON_GAMES - (r["season_2025_games"] or 0))
            r["missed_games_label"] = missed
        r["blurb"] = r["blurb"] or ""  # blank unless a blurb was already loaded previously

    return top_n


def write_position_sheet(ws, rows: list[dict]) -> None:
    for col_idx, (header, _key, width, _is_decimal) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    notes_col_idx = [i for i, (h, *_rest) in enumerate(COLUMNS, start=1) if h == "Notes"][0]
    gsis_col_idx = [i for i, (h, *_rest) in enumerate(COLUMNS, start=1) if h == "gsis_id"][0]

    for row_idx, r in enumerate(rows, start=2):
        for col_idx, (_header, key, _width, is_decimal) in enumerate(COLUMNS, start=1):
            value = r.get(key)
            if is_decimal and value is not None:
                value = round(value, 1)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BOLD_FONT if key == "full_name" else BODY_FONT
            if is_decimal:
                cell.number_format = DECIMAL_FORMAT
            if col_idx == notes_col_idx:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    last_row = len(rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"
    ws.column_dimensions[get_column_letter(gsis_col_idx)].hidden = True


def main() -> None:
    conn = sqlite3.connect(DB_PATH)
    try:
        all_rows = fetch_rankings(conn)
        adp_lookup = fetch_adp_lookup(conn)
    finally:
        conn.close()

    if not all_rows:
        raise SystemExit(f"No rankings found for scoring_format='{SCORING_FORMAT}' -- run scoring.py first.")

    wb = Workbook()
    first = True
    for position, limit in POSITION_LIMITS.items():
        rows = build_position_rows(all_rows, adp_lookup, position, limit)
        if not rows:
            print(f"No rows for position={position} -- skipping tab.")
            continue
        ws = wb.active if first else wb.create_sheet(position)
        if first:
            ws.title = position
            first = False
        write_position_sheet(ws, rows)
        print(f"{position}: wrote {len(rows)} rows")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved worklist to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
