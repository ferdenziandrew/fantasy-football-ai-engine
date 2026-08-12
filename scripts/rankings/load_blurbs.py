"""
Loads approved blurbs from the filled-in worklist files (blurb_worklist.py's output,
reviewed/edited by Andrew) into rankings.blurb / rankings.blurb_source -- the last step
in the QB/TE/RB/WR blurb-writing workflow described in blurb_worklist.py's docstring.

Matches rows by the hidden gsis_id column, not by player name (see ROADMAP.md's
"Mike Washington Jr." duplicate-name note for why name-matching isn't safe).

Every blurb loaded by this pass was Andrew-authored (no Claude-drafted blurbs went out
in this round -- the QB/TE pass left 6 players deliberately blank rather than having
Claude draft them, per Andrew's call), so blurb_source is set to 'andrew' for every row
with non-blank Notes. Blank Notes cells are skipped entirely -- not overwritten with an
empty string, since a blank cell here means "not written yet," not "confirmed no blurb."

If a future round mixes in Claude-drafted blurbs, this script would need a way to tell
the two apart (e.g. a marker in the Notes text, or a separate column) -- not needed yet,
so not built until it's an actual requirement.

Usage:
    py scripts/rankings/load_blurbs.py
"""

import sqlite3
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "data" / "db" / "fantasy_football.db"
CONTENT_DIR = PROJECT_ROOT / "content"

SCORING_FORMAT = "ppr"
BLURB_SOURCE = "andrew"  # every blurb loaded by this script this round was Andrew-authored

# Every worklist file to pull from -- add new files here as more position passes ship.
WORKLIST_FILES = [
    CONTENT_DIR / "blurb_worklist.xlsx",       # QB, TE
    CONTENT_DIR / "blurb_worklist_rbwr.xlsx",  # RB, WR
]


def extract_blurbs(path: Path) -> list[dict]:
    """Returns [{gsis_id, full_name, blurb}, ...] for every row with non-blank Notes,
    across every sheet in the workbook."""
    if not path.exists():
        print(f"  (skipping -- not found: {path})")
        return []

    wb = load_workbook(path, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        if "Notes" not in headers or "gsis_id" not in headers:
            print(f"  (skipping sheet '{sheet_name}' in {path.name} -- missing expected columns)")
            continue
        notes_col = headers.index("Notes") + 1
        gsis_col = headers.index("gsis_id") + 1
        name_col = headers.index("Player") + 1 if "Player" in headers else None

        for row in range(2, ws.max_row + 1):
            blurb = ws.cell(row=row, column=notes_col).value
            gsis_id = ws.cell(row=row, column=gsis_col).value
            if not blurb or not str(blurb).strip():
                continue  # blank -- not written yet, don't touch the DB row
            if not gsis_id:
                full_name = ws.cell(row=row, column=name_col).value if name_col else "?"
                print(f"  WARNING: row for '{full_name}' in sheet '{sheet_name}' has a blurb "
                      f"but no gsis_id -- skipping, can't match it safely")
                continue
            rows.append({
                "gsis_id": gsis_id,
                "full_name": ws.cell(row=row, column=name_col).value if name_col else None,
                "blurb": str(blurb).strip(),
            })
    return rows


def load_into_db(conn: sqlite3.Connection, rows: list[dict]) -> dict:
    """UPDATEs existing rankings rows (never INSERTs -- a blurb only makes sense for a
    player scoring.py already ranked). Returns counts for a summary."""
    updated, missing = 0, []
    for r in rows:
        cur = conn.execute(
            "UPDATE rankings SET blurb = ?, blurb_source = ? WHERE gsis_id = ? AND scoring_format = ?",
            (r["blurb"], BLURB_SOURCE, r["gsis_id"], SCORING_FORMAT),
        )
        if cur.rowcount == 0:
            missing.append(r["full_name"] or r["gsis_id"])
        else:
            updated += 1
    return {"updated": updated, "missing": missing}


def main() -> None:
    all_rows = []
    for path in WORKLIST_FILES:
        print(f"Reading {path.name}...")
        rows = extract_blurbs(path)
        print(f"  {len(rows)} non-blank blurbs found")
        all_rows.extend(rows)

    if not all_rows:
        raise SystemExit("No blurbs found across any worklist file -- nothing to load.")

    # A gsis_id could appear in more than one file if a worklist gets re-run and
    # re-filled -- last one wins, so later files in WORKLIST_FILES take precedence.
    by_gsis_id = {r["gsis_id"]: r for r in all_rows}
    print(f"\n{len(by_gsis_id)} unique players with a blurb across all files")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        result = load_into_db(conn, list(by_gsis_id.values()))
        conn.commit()
    finally:
        conn.close()

    print(f"\nUpdated {result['updated']} rankings rows (scoring_format='{SCORING_FORMAT}').")
    if result["missing"]:
        print(f"WARNING: {len(result['missing'])} players had a blurb but no matching "
              f"rankings row for scoring_format='{SCORING_FORMAT}' -- not loaded:")
        for name in result["missing"][:20]:
            print(f"    {name}")


if __name__ == "__main__":
    main()
