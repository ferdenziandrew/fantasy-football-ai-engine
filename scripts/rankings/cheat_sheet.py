"""
Builds a formatted Excel cheat sheet from the rankings table: one "Overall" tab (all
players, sorted by overall value-over-replacement rank) plus one tab per position
(QB/RB/WR/TE/K) for the drill-down view, in that fixed order.

Design history:
- v1 used a full-width merged row between groups ("TIER 5", "PICKS 11-20"). Broke in
  practice -- Excel refuses to sort/filter a range with merged cells of inconsistent
  size. Replaced with a plain "Tier" data column instead.
- v2 used STATIC alternating row fills tied to tier group order at generation time.
  Didn't reflow correctly once sorted by a different column in Excel. Replaced with
  conditional formatting keyed to ROW POSITION (`=MOD(ROW(),2)=0`).
- v3 dropped the separate "Pos" column -- redundant with "Pos Rank" and the color-coding.
- v4 (2026-07-28): row-position striping read as "too white" and didn't visually show
  tier boundaries at all. Switched to conditional formatting keyed to the TIER VALUE
  itself (ISEVEN/ISODD on the Tier column, now a plain integer instead of "Tier N" text)
  -- shows real tier boundaries and still survives sorting, since a row's own Tier
  value travels with it. Loosened the tier-gap threshold since 1.5x was too granular.
- v5 (2026-07-28): column layout now differs by tab. QB cares about passing stats
  first; every other position (and the Overall tab, since QBs are a minority of its
  rows) cares about rush/rec stats first and passing last, since it's almost always
  zero/irrelevant there. "2025 PPG" moved right after Tier -- the stat Andrew's
  actually cross-referencing against during this tuning pass.
- v6 (2026-07-28): tier/rookie coloring still wasn't rendering after v4/v5 -- root
  cause was that PatternFill objects used inside conditional formatting need both
  start_color and end_color set (see the comment above ROOKIE_FILL/TIER_FILL_A/B).
  A regular static PatternFill (fgColor only, the way POSITION_FILLS is built) is
  invisible when used as a conditional-formatting fill instead of a cell fill.
- v7 (2026-07-28): added per-game rate columns next to Rush Yds, Pass Yds, and
  Receptions (Rush YPG, Pass YPG, Rec/Gm) -- bounded to yardage/reception counts, not
  every stat, since TD/game is too noisy to be a useful signal at these counts and
  doubling every column would clutter the sheet past the point of being a quick
  reference. Floor/Ceiling moved to just before Rookie/Notes (Andrew's ask -- wanted
  them out of the way of the stats he's actively cross-referencing near the front).
- v8 (2026-07-28): considered combining total + rate into one text cell ("3668 /
  229.2", Andrew's original phrasing) but that sacrifices numeric sort/filter on that
  column (Excel would sort it as text) -- Andrew's call after seeing the tradeoff was
  to keep them as separate real-number columns. Added two more rate columns (Rush
  Att/Gm, Rec YPG) alongside the existing three. Separately, the AutoFilter dropdown
  arrow was visually colliding with header text on narrow columns -- fixed by
  left-aligning headers (text hugs the left edge, arrow has clean room on the right)
  and widening the narrower numeric columns a few characters.

Because column order/count now varies by tab, all the "which column is X" lookups
(player name, Pos Rank, Tier, Rookie) are computed per-call from that sheet's own
column list, not as fixed module-level constants.

Raw 2025 season stats (and 2025 PPG = season points / season games) come from the
`season_stats` VIEW, derived live from weekly_stats -- no new ingestion needed. Rush+Rec
Yds is computed here (rushing + receiving yards), not a stored column.

Note on "highlight the column I'm sorting by": clicking a column's letter header
already does this natively in Excel. Making it automatic on every sort would need VBA
macros (.xlsm, with the macro-security prompts that come with it) -- not done here.

Blurbs (column header "Notes") are left blank for now -- a separate pass (Andrew writes
some, Claude drafts the rest, Andrew approves) planned for after this structure is
confirmed useful.

Usage:
    py scripts/rankings/cheat_sheet.py
"""

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "data" / "db" / "fantasy_football.db"
OUTPUT_PATH = PROJECT_ROOT / "content" / "cheat_sheet.xlsx"

SCORING_FORMAT = "ppr"
POSITION_ORDER = ["QB", "RB", "WR", "TE", "K"]
OVERALL_BAND_SIZE = 10  # rows per pick-band group on the Overall tab
TIER_GAP_MULTIPLIER = 2.0  # position tabs: a gap this many times the average = new tier
# (started at 1.5, which was too granular on real data; 3.0 in testing swung too far the
# other way on a synthetic approximation of the real curve -- 2.0 is a reasonable middle
# ground, but may need one more real-data adjustment once you see the actual tier counts.)
DECIMAL_FORMAT = "0.0"
SEASON_STATS_YEAR = 2025  # "last year's stats" -- update each offseason

BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="404040")
WHITE_BOLD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")

# NOTE on these three fills specifically: they're used inside FormulaRule (conditional
# formatting), not as static cell fills. Excel's differential-formatting (dxf) renderer
# reads a fill's BACKGROUND color for a solid pattern, not the foreground -- the reverse
# of how a normal static cell fill works, where fgColor is what's visible (see
# POSITION_FILLS below, which is correct as-is since those are static). A PatternFill
# built the normal way (fgColor only) silently renders as no color at all inside a
# FormulaRule -- the rule and color data are genuinely in the file, they just never
# show. Fix: set start_color AND end_color to the same hex, so it's correct either way
# Excel's dxf renderer decides to read it. (This likely also explains why the very
# first zebra-striping attempt read as "all white" -- it probably wasn't rendering at
# all, not just rendering too subtly.)
ROOKIE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light gold

# Two clearly distinct, non-neon pastel tones -- alternated by TIER VALUE (not row
# position), so a color change on screen means "new tier," not just "next row."
TIER_FILL_A = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # soft blue
TIER_FILL_B = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # soft peach

# Dark, distinct, white-text-readable tones -- used only on the Overall tab's Pos Rank
# column, since position tabs are already single-position and don't need this.
POSITION_FILLS = {
    "QB": PatternFill("solid", fgColor="7030A0"),  # purple
    "RB": PatternFill("solid", fgColor="385723"),  # dark green
    "WR": PatternFill("solid", fgColor="1F4E78"),  # dark blue
    "TE": PatternFill("solid", fgColor="BF8F00"),  # dark gold
    "K":  PatternFill("solid", fgColor="632423"),  # dark maroon
}

# Shared prefix/suffix for every tab; only the stat block in between changes by tab.
# Floor/Ceiling live in the suffix (just before Rookie/Notes), not here -- Andrew's
# ask (2026-07-28) was to get them out of the way of the stats he's actively
# cross-referencing near the front of the sheet.
# Column widths below are a few characters wider than the header text needs on its
# own -- AutoFilter's dropdown arrow sits inside the header cell's right edge, and on
# a tightly-sized column it visually collides with the last letter or two of the
# header. The extra width is clearance for that arrow, not a typo.
BASE_PREFIX = [
    ("Rank", "rank", 8),
    ("Player", "full_name", 25),
    ("Pos Rank", "pos_rank_label", 11),
    ("Team", "current_team", 8),
    ("Tier", "tier_num", 8),
    ("2025 PPG", "ppg_2025", 11),
    ("Score", "score", 10),
    ("Games", "games_played", 10),
]
PASSING_COLS = [
    ("Pass Yds", "passing_yards", 11),
    ("Pass YPG", "pass_ypg", 11),
    ("Pass TD", "passing_tds", 10),
]
RUSH_REC_COLS = [
    ("Rush Att", "rush_attempts", 11),
    ("Rush Att/Gm", "rush_att_pg", 13),
    ("Rush Yds", "rushing_yards", 11),
    ("Rush YPG", "rush_ypg", 11),
    ("Rush TD", "rushing_tds", 10),
    ("Targets", "targets", 10),
    ("Rec", "receptions", 8),
    ("Rec/Gm", "rec_per_game", 10),
    ("Rec Yds", "receiving_yards", 10),
    ("Rec YPG", "rec_ypg", 10),
    ("Rec TD", "receiving_tds", 9),
    ("R+R Yds", "rush_rec_yards", 10),
]
SUFFIX = [
    ("Floor", "floor", 10),
    ("Ceiling", "ceiling", 11),
    ("Rookie", "is_rookie_baseline", 11),
    ("Notes", "blurb", 40),
]

DECIMAL_KEYS = {
    "score", "floor", "ceiling", "ppg_2025",
    "pass_ypg", "rush_ypg", "rec_per_game", "rush_att_pg", "rec_ypg",
}
SEASON_STAT_KEYS = {
    "passing_yards", "passing_tds", "rush_attempts", "rushing_yards", "rushing_tds",
    "targets", "receptions", "receiving_yards", "receiving_tds", "rush_rec_yards",
}


def get_columns(is_qb: bool) -> list:
    """QB cares about passing stats first; every other tab (including Overall, since
    QBs are a minority of its rows) cares about rush/rec first and passing last, since
    it's almost always zero/irrelevant there."""
    stat_cols = (PASSING_COLS + RUSH_REC_COLS) if is_qb else (RUSH_REC_COLS + PASSING_COLS)
    return BASE_PREFIX + stat_cols + SUFFIX


def col_idx_for(columns: list, header: str) -> int:
    return [i for i, (h, _k, _w) in enumerate(columns, start=1) if h == header][0]


def fetch_rankings(conn: sqlite3.Connection) -> list[dict]:
    cur = conn.execute(
        """
        SELECT r.gsis_id, r.rank, p.full_name, p.position, p.current_team, r.positional_rank,
               r.score, r.games_played, r.floor, r.ceiling, r.is_rookie_baseline, r.blurb,
               s.passing_yards, s.passing_tds, s.carries AS rush_attempts,
               s.rushing_yards, s.rushing_tds, s.targets, s.receptions,
               s.receiving_yards, s.receiving_tds,
               s.fantasy_points_ppr AS season_2025_points, s.games_played AS season_2025_games
        FROM rankings r
        JOIN players p ON r.gsis_id = p.gsis_id
        LEFT JOIN season_stats s ON s.gsis_id = r.gsis_id AND s.season = ?
        WHERE r.scoring_format = ?
        ORDER BY r.rank
        """,
        (SEASON_STATS_YEAR, SCORING_FORMAT),
    )
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, row)) for row in cur.fetchall()]
    for r in rows:
        r["pos_rank_label"] = f"{r['position']}{r['positional_rank']}"
        for key in SEASON_STAT_KEYS - {"rush_rec_yards"}:
            r[key] = r[key] or 0
        r["rush_rec_yards"] = r["rushing_yards"] + r["receiving_yards"]
        games = r["season_2025_games"]
        r["ppg_2025"] = (r["season_2025_points"] or 0) / games if games else 0
        # Per-game rate stats -- bounded to yardage/attempt/reception counts (not every
        # stat, e.g. not TDs -- too noisy at these counts). Same games-played
        # denominator as 2025 PPG, so all the rate columns on a row are comparable.
        r["pass_ypg"] = r["passing_yards"] / games if games else 0
        r["rush_ypg"] = r["rushing_yards"] / games if games else 0
        r["rush_att_pg"] = r["rush_attempts"] / games if games else 0
        r["rec_per_game"] = r["receptions"] / games if games else 0
        r["rec_ypg"] = r["receiving_yards"] / games if games else 0
    return rows


def compute_score_gap_tiers(rows_sorted_by_score_desc: list[dict]) -> list[int]:
    """Assigns a tier index per row based on natural gaps in score, not a fixed group
    size -- a new tier starts where there's a real drop-off in value at this position,
    not an arbitrary row count."""
    if len(rows_sorted_by_score_desc) <= 1:
        return [0] * len(rows_sorted_by_score_desc)

    scores = [r["score"] for r in rows_sorted_by_score_desc]
    gaps = [scores[i] - scores[i + 1] for i in range(len(scores) - 1)]
    avg_gap = sum(gaps) / len(gaps) if gaps else 0
    threshold = max(avg_gap * TIER_GAP_MULTIPLIER, 0.1)  # floor avoids noise from tiny/float gaps

    tiers = [0]
    current_tier = 0
    for gap in gaps:
        if gap > threshold:
            current_tier += 1
        tiers.append(current_tier)
    return tiers


def write_data_row(ws, row_idx: int, r: dict, columns: list, player_col_idx: int,
                    pos_rank_col_idx: int, color_code_positions: bool) -> None:
    position_fill = POSITION_FILLS.get(r["position"])

    for col_idx, (_header, key, _width) in enumerate(columns, start=1):
        value = r[key]
        if key == "is_rookie_baseline":
            value = "Yes" if value else ""
        elif key in DECIMAL_KEYS and value is not None:
            value = round(value, 1)

        cell = ws.cell(row=row_idx, column=col_idx, value=value)

        if color_code_positions and col_idx == pos_rank_col_idx and position_fill:
            cell.fill = position_fill
            cell.font = WHITE_BOLD_FONT
        else:
            cell.font = BOLD_FONT if col_idx == player_col_idx else BODY_FONT

        if key in DECIMAL_KEYS:
            cell.number_format = DECIMAL_FORMAT


def apply_conditional_formatting(ws, last_row: int, num_cols: int, pos_rank_col_idx: int,
                                  tier_col_idx: int, rookie_col_idx: int, color_code_positions: bool) -> None:
    """Tier-based row coloring and rookie highlighting as live Excel rules, so they
    read off actual cell data (the Tier/Rookie columns) rather than row position --
    this means the coloring survives any sort/filter, since a row's Tier/Rookie value
    travels with it. Skips the Pos Rank column on the Overall tab so the position
    color-coding always shows through."""
    last_col_letter = get_column_letter(num_cols)
    tier_col_letter = get_column_letter(tier_col_idx)
    rookie_col_letter = get_column_letter(rookie_col_idx)

    if color_code_positions:
        ranges = [f"A2:{get_column_letter(pos_rank_col_idx - 1)}{last_row}",
                  f"{get_column_letter(pos_rank_col_idx + 1)}2:{last_col_letter}{last_row}"]
    else:
        ranges = [f"A2:{last_col_letter}{last_row}"]

    for rng in ranges:
        # Rookie rule first with stopIfTrue -- a rookie row shows gold and the tier
        # coloring below is skipped for it, rather than painting over the gold.
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${rookie_col_letter}2="Yes"'], fill=ROOKIE_FILL, stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f"ISEVEN(${tier_col_letter}2)"], fill=TIER_FILL_A),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f"ISODD(${tier_col_letter}2)"], fill=TIER_FILL_B),
        )


def write_sheet(ws, rows: list[dict], sort_key: str, band_mode: str, is_qb: bool) -> None:
    rows = sorted(rows, key=lambda r: r[sort_key])
    color_code_positions = band_mode == "fixed_rows"  # Overall tab only
    columns = get_columns(is_qb)
    num_cols = len(columns)
    player_col_idx = col_idx_for(columns, "Player")
    pos_rank_col_idx = col_idx_for(columns, "Pos Rank")
    tier_col_idx = col_idx_for(columns, "Tier")
    rookie_col_idx = col_idx_for(columns, "Rookie")

    if band_mode == "score_gap":
        band_indices = compute_score_gap_tiers(rows)
    else:  # "fixed_rows" -- Overall tab, grouped by row position (already sorted by overall rank)
        band_indices = [i // OVERALL_BAND_SIZE for i in range(len(rows))]

    for r, band_idx in zip(rows, band_indices):
        r["tier_num"] = band_idx + 1  # 1-indexed for display -- "Tier 0" reads oddly

    for col_idx, (header, _key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        # Left-aligned, not centered: the AutoFilter dropdown arrow occupies the right
        # edge of the header cell, so left-aligned text keeps clear of it -- centered
        # text tends to run into the arrow on narrower columns.
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows):
        write_data_row(ws, i + 2, r, columns, player_col_idx, pos_rank_col_idx, color_code_positions)

    last_row = len(rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{last_row}"
    apply_conditional_formatting(ws, last_row, num_cols, pos_rank_col_idx, tier_col_idx,
                                  rookie_col_idx, color_code_positions)


def write_legend(ws) -> None:
    ws.column_dimensions["A"].width = 90
    lines = [
        "Legend",
        "",
        "Score: value over replacement (VOR) -- how much better this player is than the last",
        "  realistically-startable player at their position, not raw fantasy points.",
        "Pos Rank: rank within the player's own position, prefixed with position (RB1, RB2, etc.)",
        "  -- no separate Pos column since this already conveys it (and color does too, on Overall).",
        "Tier (position tabs): players roughly interchangeable in value -- a new tier means a real",
        f"  gap in score, not a fixed group size. Tier (Overall tab): every {OVERALL_BAND_SIZE} picks,",
        "  roughly draft-round-sized chunks -- not tied to value gaps, just a readability grouping.",
        "  Row color alternates by tier (blue/peach), not by row position, so it stays meaningful",
        "  even if you sort by a different column.",
        f"2025 PPG: full-PPR points per game in the {SEASON_STATS_YEAR} season specifically (0 if they",
        "  didn't play that season -- rookies, or missed the whole year).",
        "Games: number of games in the historical window used to compute the score",
        f"Pass/Rush/Rec stats: full {SEASON_STATS_YEAR} season totals. R+R Yds = rushing + receiving",
        "  yards combined (useful for pass-catching RBs). QB tab shows passing stats first; every",
        "  other tab (including Overall) shows rush/rec first since passing is usually 0 there.",
        "Pass YPG / Rush YPG / Rec YPG / Rush Att/Gm / Rec/Gm: per-game rate versions of the",
        "  adjacent total column, using the same 2025 games-played denominator as 2025 PPG. TDs",
        "  are totals only, not rates -- too noisy as a per-game number at these counts to matter.",
        "Floor / Ceiling: 25th / 75th percentile of this player's own weekly scores (recent history)",
        "Rookie (highlighted gold): this year's draft class with zero games played -- score is a",
        "  baseline derived from draft position only, not actual performance. Treat with more",
        "  caution than a performance-based score.",
        "Pos Rank colors (Overall tab only): QB purple, RB green, WR blue, TE gold, K maroon.",
        "Every column is filterable/sortable -- use the dropdown arrows on the header row. Clicking a",
        "  column's letter (above the header row) selects/highlights that whole column.",
        "Notes: left blank for now -- a future pass adds player notes here.",
        "",
        f"Generated from scoring_format='{SCORING_FORMAT}'. Re-run scoring.py after changing",
        "weights_config.py, then re-run this script to refresh the sheet.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(name="Arial", size=10, bold=(i == 1))


def main() -> None:
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = fetch_rankings(conn)
    finally:
        conn.close()

    if not rows:
        raise SystemExit(f"No rankings found for scoring_format='{SCORING_FORMAT}' -- run scoring.py first.")

    wb = Workbook()

    overall = wb.active
    overall.title = "Overall"
    write_sheet(overall, rows, sort_key="rank", band_mode="fixed_rows", is_qb=False)

    available_positions = {r["position"] for r in rows}
    for position in POSITION_ORDER:
        if position not in available_positions:
            continue
        ws = wb.create_sheet(position)
        position_rows = [r for r in rows if r["position"] == position]
        write_sheet(ws, position_rows, sort_key="positional_rank", band_mode="score_gap", is_qb=(position == "QB"))

    legend = wb.create_sheet("Legend")
    write_legend(legend)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {len(rows)} players across {len(available_positions)} position tabs to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
